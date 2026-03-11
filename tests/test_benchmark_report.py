from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
SCRIPTS_DIR = REPO_ROOT / "external" / "carts-benchmarks" / "scripts"
TOOLS_DIR = REPO_ROOT / "tools"

sys.path.insert(0, str(TOOLS_DIR))
sys.path.insert(0, str(SCRIPTS_DIR))

from benchmark_models import ExperimentStep  # noqa: E402
from benchmark_report import generate_report_from_rows  # noqa: E402


class BenchmarkReportWorkbookTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tmpdir.name)
        self.experiment_dir = self.root / "exp_report"
        self.experiment_dir.mkdir(parents=True, exist_ok=True)

    def tearDown(self) -> None:
        self.tmpdir.cleanup()

    def _remoteize(self, local_path: Path) -> str:
        rel = local_path.relative_to(self.experiment_dir)
        return str(Path("/remote/results") / self.experiment_dir.name / rel)

    def _write_counter_bundle(
        self,
        counter_dir: Path,
        *,
        total_nodes: int,
        cluster_counters: dict[str, dict[str, object]],
        node_counters: list[dict[str, dict[str, object]]],
    ) -> None:
        counter_dir.mkdir(parents=True, exist_ok=True)
        (counter_dir / "cluster.json").write_text(json.dumps({"counters": cluster_counters}))
        for node_id, counters in enumerate(node_counters):
            payload = {
                "metadata": {
                    "nodeId": node_id,
                    "totalThreads": 64,
                    "totalNodes": total_nodes,
                },
                "counters": counters,
            }
            (counter_dir / f"n{node_id}.json").write_text(json.dumps(payload))

    def _make_artifacts(self, phase: str, *, threads: int, nodes: int) -> dict[str, str]:
        run_dir = (
            self.experiment_dir
            / phase
            / "polybench/gemm"
            / f"{threads}t_{nodes}n"
            / "run_1"
        )
        artifact_dir = run_dir.parent / "artifacts"
        counter_dir = run_dir / "counters"
        artifact_dir.mkdir(parents=True, exist_ok=True)
        run_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "arts.cfg").write_text("threads=64\n")
        (artifact_dir / "build.log").write_text("ok\n")
        (run_dir / "run_config.json").write_text("{}\n")
        (run_dir / "result.json").write_text("{}\n")
        (run_dir / "slurm.out").write_text("")
        (run_dir / "slurm.err").write_text("")
        return {
            "run_dir": self._remoteize(run_dir),
            "run_config": self._remoteize(run_dir / "run_config.json"),
            "result_json": self._remoteize(run_dir / "result.json"),
            "slurm_out": self._remoteize(run_dir / "slurm.out"),
            "slurm_err": self._remoteize(run_dir / "slurm.err"),
            "build_dir": self._remoteize(artifact_dir),
            "arts_config": self._remoteize(artifact_dir / "arts.cfg"),
            "counter_dir": self._remoteize(counter_dir),
        }

    def _make_result(
        self,
        *,
        phase: str,
        size: str,
        threads: int,
        nodes: int,
        arts_e2e: float,
        omp_e2e: float | None,
        compile_args: str | None = None,
        artifacts: dict[str, str],
    ) -> dict[str, object]:
        omp_payload: dict[str, object]
        if omp_e2e is None:
            omp_payload = {"exit_code": 0, "duration_sec": 0.0, "checksum": None, "skipped": True}
        else:
            omp_payload = {
                "exit_code": 0,
                "duration_sec": omp_e2e,
                "checksum": "1.0",
                "e2e_timings": {"gemm": omp_e2e},
                "kernel_timings": {},
                "skipped": False,
            }
        verification: dict[str, object] = {"note": "Checksums match"}
        if omp_e2e is None:
            verification["reference_checksum"] = "1.0"
            verification["reference_source"] = self._remoteize(
                self.experiment_dir / "references" / phase / f"{threads}t_{nodes}n" / "reference.json"
            )
            verification["reference_omp_threads"] = threads
            verification["mode"] = "stored_omp_reference"
        else:
            verification["mode"] = "direct_omp"
        return {
            "benchmark": "polybench/gemm",
            "size": size,
            "threads": threads,
            "nodes": nodes,
            "run_number": 1,
            "run_phase": phase,
            "compile_args": compile_args,
            "profile": "profile-overhead.cfg" if "overhead" in phase else None,
            "perf": False,
            "perf_interval": None,
            "status": "PASS",
            "status_detail": "PASS",
            "verification": verification,
            "arts": {
                "exit_code": 0,
                "duration_sec": arts_e2e,
                "checksum": "1.0",
                "e2e_timings": {"gemm": arts_e2e},
                "kernel_timings": {},
            },
            "omp": omp_payload,
            "slurm": {
                "job_id": f"{phase}-{threads}-{nodes}",
                "state": "COMPLETED",
                "exit_code": 0,
                "nodelist": "j001",
            },
            "diagnostics": {},
            "artifacts": artifacts,
        }

    def test_generate_report_from_rows_builds_analysis_sheets(self) -> None:
        thread_1_artifacts = self._make_artifacts("thread-sweep", threads=1, nodes=1)
        thread_2_artifacts = self._make_artifacts("thread-sweep", threads=2, nodes=1)
        base_2_artifacts = self._make_artifacts("multinode-overhead-baseline", threads=64, nodes=2)
        base_4_artifacts = self._make_artifacts("multinode-overhead-baseline", threads=64, nodes=4)
        dist_2_artifacts = self._make_artifacts(
            "multinode-overhead-distributed-db", threads=64, nodes=2
        )
        dist_4_artifacts = self._make_artifacts(
            "multinode-overhead-distributed-db", threads=64, nodes=4
        )

        self._write_counter_bundle(
            self.experiment_dir
            / "multinode-overhead-baseline/polybench/gemm/64t_2n/run_1/counters",
            total_nodes=2,
            cluster_counters={
                "memoryFootprint": {"value": 1000},
                "remoteBytesSent": {"value": 100},
                "remoteBytesReceived": {"value": 80},
                "numEdtsCreated": {"value": 20},
                "numDbsCreated": {"value": 10},
                "initializationTime": {"value_ms": 10.0},
                "endToEndTime": {"value_ms": 100000.0},
            },
            node_counters=[
                {
                    "memoryFootprint": {"captureLevel": "CLUSTER", "value": 700},
                    "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 90},
                    "remoteBytesReceived": {"captureLevel": "CLUSTER", "value": 50},
                },
                {
                    "memoryFootprint": {"captureLevel": "CLUSTER", "value": 300},
                    "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 10},
                    "remoteBytesReceived": {"captureLevel": "CLUSTER", "value": 30},
                },
            ],
        )
        self._write_counter_bundle(
            self.experiment_dir
            / "multinode-overhead-baseline/polybench/gemm/64t_4n/run_1/counters",
            total_nodes=4,
            cluster_counters={
                "memoryFootprint": {"value": 1600},
                "remoteBytesSent": {"value": 400},
                "remoteBytesReceived": {"value": 380},
                "numEdtsCreated": {"value": 40},
                "numDbsCreated": {"value": 18},
                "initializationTime": {"value_ms": 11.0},
                "endToEndTime": {"value_ms": 70000.0},
            },
            node_counters=[
                {"memoryFootprint": {"captureLevel": "CLUSTER", "value": 900}, "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 280}},
                {"memoryFootprint": {"captureLevel": "CLUSTER", "value": 300}, "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 60}},
                {"memoryFootprint": {"captureLevel": "CLUSTER", "value": 200}, "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 40}},
                {"memoryFootprint": {"captureLevel": "CLUSTER", "value": 200}, "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 20}},
            ],
        )
        self._write_counter_bundle(
            self.experiment_dir
            / "multinode-overhead-distributed-db/polybench/gemm/64t_2n/run_1/counters",
            total_nodes=2,
            cluster_counters={
                "memoryFootprint": {"value": 950},
                "remoteBytesSent": {"value": 120},
                "remoteBytesReceived": {"value": 120},
                "numEdtsCreated": {"value": 28},
                "numDbsCreated": {"value": 10},
                "initializationTime": {"value_ms": 14.0},
                "endToEndTime": {"value_ms": 110000.0},
            },
            node_counters=[
                {
                    "memoryFootprint": {"captureLevel": "CLUSTER", "value": 520},
                    "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 70},
                    "remoteBytesReceived": {"captureLevel": "CLUSTER", "value": 60},
                },
                {
                    "memoryFootprint": {"captureLevel": "CLUSTER", "value": 430},
                    "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 50},
                    "remoteBytesReceived": {"captureLevel": "CLUSTER", "value": 60},
                },
            ],
        )
        self._write_counter_bundle(
            self.experiment_dir
            / "multinode-overhead-distributed-db/polybench/gemm/64t_4n/run_1/counters",
            total_nodes=4,
            cluster_counters={
                "memoryFootprint": {"value": 1400},
                "remoteBytesSent": {"value": 320},
                "remoteBytesReceived": {"value": 330},
                "numEdtsCreated": {"value": 52},
                "numDbsCreated": {"value": 18},
                "initializationTime": {"value_ms": 16.0},
                "endToEndTime": {"value_ms": 50000.0},
            },
            node_counters=[
                {"memoryFootprint": {"captureLevel": "CLUSTER", "value": 380}, "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 90}},
                {"memoryFootprint": {"captureLevel": "CLUSTER", "value": 340}, "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 80}},
                {"memoryFootprint": {"captureLevel": "CLUSTER", "value": 340}, "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 75}},
                {"memoryFootprint": {"captureLevel": "CLUSTER", "value": 340}, "remoteBytesSent": {"captureLevel": "CLUSTER", "value": 75}},
            ],
        )

        results = [
            self._make_result(
                phase="thread-sweep",
                size="large",
                threads=1,
                nodes=1,
                arts_e2e=10.0,
                omp_e2e=20.0,
                artifacts=thread_1_artifacts,
            ),
            self._make_result(
                phase="thread-sweep",
                size="large",
                threads=2,
                nodes=1,
                arts_e2e=5.0,
                omp_e2e=15.0,
                artifacts=thread_2_artifacts,
            ),
            self._make_result(
                phase="multinode-overhead-baseline",
                size="extralarge",
                threads=64,
                nodes=2,
                arts_e2e=100.0,
                omp_e2e=None,
                artifacts=base_2_artifacts,
            ),
            self._make_result(
                phase="multinode-overhead-baseline",
                size="extralarge",
                threads=64,
                nodes=4,
                arts_e2e=70.0,
                omp_e2e=None,
                artifacts=base_4_artifacts,
            ),
            self._make_result(
                phase="multinode-overhead-distributed-db",
                size="extralarge",
                threads=64,
                nodes=2,
                arts_e2e=110.0,
                omp_e2e=None,
                compile_args="--distributed-db",
                artifacts=dist_2_artifacts,
            ),
            self._make_result(
                phase="multinode-overhead-distributed-db",
                size="extralarge",
                threads=64,
                nodes=4,
                arts_e2e=50.0,
                omp_e2e=None,
                compile_args="--distributed-db",
                artifacts=dist_4_artifacts,
            ),
        ]

        steps = [
            ExperimentStep(
                name="thread-sweep",
                description="Single-node scaling study",
                size="large",
                threads="1,2",
                nodes="1",
                runs=5,
            ),
            ExperimentStep(
                name="multinode-overhead-baseline",
                description="Baseline multinode diagnostics",
                size="extralarge",
                threads="64",
                nodes="2,4",
                runs=1,
                profile="profile-overhead.cfg",
            ),
            ExperimentStep(
                name="multinode-overhead-distributed-db",
                description="Distributed DB multinode diagnostics",
                size="extralarge",
                threads="64",
                nodes="2,4",
                runs=1,
                profile="profile-overhead.cfg",
                compile_args="--distributed-db",
            ),
        ]
        setattr(steps[0], "_experiment_name", "report-fixture")
        setattr(steps[0], "_experiment_description", "Synthetic workbook regression fixture")

        report_path = generate_report_from_rows(results, self.experiment_dir, steps=steps)
        self.assertIsNotNone(report_path)
        workbook = load_workbook(report_path, data_only=False)

        self.assertEqual(
            workbook.sheetnames,
            [
                "Guide",
                "Overview",
                "Summary",
                "ThreadScaling",
                "NodeScaling",
                "DistributedDbDelta",
                "ScalingMatrix",
                "Comparison",
                "NodeCounterSummary",
                "NodeCounters",
                "Results",
                "Metadata",
            ],
        )

        self.assertEqual(len(workbook["Guide"].tables), 1)
        self.assertEqual(len(workbook["Results"].tables), 1)

        guide_rows = list(workbook["Guide"].iter_rows(values_only=True))
        self.assertEqual(guide_rows[1][0], "Overview")
        self.assertEqual(guide_rows[1][3], '=HYPERLINK("#\'Overview\'!A1","Open")')
        self.assertEqual(workbook["Guide"]["F4"].value, "Experiment")
        self.assertEqual(workbook["Guide"]["F5"].value, "report-fixture")
        self.assertEqual(
            workbook["Guide"]["F6"].value,
            "Synthetic workbook regression fixture",
        )
        self.assertEqual(workbook["Guide"].column_dimensions["F"].width, 64.0)
        self.assertEqual(workbook["Guide"].sheet_properties.tabColor.rgb, "004472C4")
        self.assertTrue(workbook["Guide"]["F6"].alignment.wrap_text)

        overview_rows = list(workbook["Overview"].iter_rows(values_only=True))
        overview_header = {name: idx for idx, name in enumerate(overview_rows[0])}
        thread_overview = next(row for row in overview_rows[1:] if row[overview_header["step"]] == "thread-sweep")
        self.assertEqual(
            thread_overview[overview_header["step_description"]],
            "Single-node scaling study",
        )
        self.assertEqual(thread_overview[overview_header["threads"]], "1,2")
        self.assertEqual(thread_overview[overview_header["nodes"]], "1")
        self.assertEqual(thread_overview[overview_header["rows"]], 2)
        self.assertEqual(thread_overview[overview_header["passed"]], 2)
        self.assertEqual(thread_overview[overview_header["verified"]], 2)
        self.assertEqual(workbook["Overview"].column_dimensions["B"].width, 56.0)
        self.assertEqual(workbook["Overview"].column_dimensions["K"].width, 34.0)
        self.assertEqual(workbook["Overview"].column_dimensions["L"].width, 30.0)
        self.assertEqual(workbook["Overview"].sheet_properties.tabColor.rgb, "005B9BD5")
        self.assertTrue(workbook["Overview"]["B2"].alignment.wrap_text)

        thread_scaling_rows = list(workbook["ThreadScaling"].iter_rows(values_only=True))
        thread_header = {name: idx for idx, name in enumerate(thread_scaling_rows[0])}
        thread_row = next(
            row
            for row in thread_scaling_rows[1:]
            if row[thread_header["threads"]] == 2
        )
        self.assertEqual(thread_row[thread_header["self_scaling"]], 2.0)

        node_scaling_rows = list(workbook["NodeScaling"].iter_rows(values_only=True))
        node_header = {name: idx for idx, name in enumerate(node_scaling_rows[0])}
        baseline_4n = next(
            row
            for row in node_scaling_rows[1:]
            if row[node_header["run_phase"]] == "multinode-overhead-baseline"
            and row[node_header["nodes"]] == 4
        )
        self.assertAlmostEqual(
            baseline_4n[node_header["self_scaling"]],
            100.0 / 70.0,
            places=6,
        )

        delta_rows = list(workbook["DistributedDbDelta"].iter_rows(values_only=True))
        delta_header = {name: idx for idx, name in enumerate(delta_rows[0])}
        delta_by_nodes = {row[delta_header["nodes"]]: row for row in delta_rows[1:]}
        self.assertAlmostEqual(delta_by_nodes[2][delta_header["time_ratio_dist_vs_base"]], 1.1, places=6)
        self.assertAlmostEqual(
            delta_by_nodes[4][delta_header["time_ratio_dist_vs_base"]],
            50.0 / 70.0,
            places=6,
        )
        self.assertEqual(delta_by_nodes[4][delta_header["winner"]], "distributed-db")
        self.assertTrue(
            str(delta_by_nodes[4][delta_header["distributed_counter_dir"]]).startswith(
                str(self.experiment_dir.resolve())
            )
        )

        node_summary_rows = list(workbook["NodeCounterSummary"].iter_rows(values_only=True))
        node_summary_header = {name: idx for idx, name in enumerate(node_summary_rows[0])}
        remote_rows = [
            row
            for row in node_summary_rows[1:]
            if row[node_summary_header["counter_name"]] == "remoteBytesSent"
        ]
        self.assertTrue(remote_rows)
        self.assertTrue(any(row[node_summary_header["nodes_reported"]] == 4 for row in remote_rows))

        result_header = {cell.value: idx + 1 for idx, cell in enumerate(workbook["Results"][1])}
        run_config_cell = workbook["Results"].cell(row=2, column=result_header["artifact_run_config"])
        self.assertIsNotNone(run_config_cell.hyperlink)
        self.assertEqual(workbook["Results"].cell(row=2, column=result_header["verification_mode"]).value, "direct_omp")

        distributed_result_row = next(
            row for row in workbook["Results"].iter_rows(min_row=2, values_only=True)
            if row[result_header["run_phase"] - 1] == "multinode-overhead-distributed-db"
            and row[result_header["nodes"] - 1] == 4
        )
        self.assertEqual(distributed_result_row[result_header["verification_mode"] - 1], "stored_omp_reference")
        self.assertEqual(distributed_result_row[result_header["reference_omp_threads"] - 1], 64)

    def test_generate_report_from_local_result_schema(self) -> None:
        report_path = generate_report_from_rows(
            [
                {
                    "name": "polybench/gemm",
                    "suite": "polybench",
                    "size": "small",
                    "run_phase": "single-node",
                    "config": {
                        "arts_threads": 4,
                        "arts_nodes": 1,
                        "omp_threads": 4,
                        "launcher": "local",
                    },
                    "run_number": 1,
                    "run_arts": {
                        "status": "PASS",
                        "duration_sec": 1.0,
                        "exit_code": 0,
                        "checksum": "10.0",
                        "kernel_timings": {},
                        "e2e_timings": {"gemm": 1.0},
                    },
                    "run_omp": {
                        "status": "PASS",
                        "duration_sec": 1.2,
                        "exit_code": 0,
                        "checksum": "10.0",
                        "kernel_timings": {},
                        "e2e_timings": {"gemm": 1.2},
                    },
                    "timing": {
                        "arts_time_sec": 1.0,
                        "omp_time_sec": 1.2,
                        "speedup": 1.2,
                        "speedup_basis": "e2e",
                    },
                    "verification": {
                        "correct": True,
                        "note": "Checksums match within tolerance",
                        "arts_checksum": "10.0",
                        "omp_checksum": "10.0",
                        "mode": "direct_omp",
                    },
                    "artifacts": {
                        "run_dir": str(self.experiment_dir / "single-node" / "run_1"),
                    },
                }
            ],
            self.experiment_dir,
        )
        self.assertIsNotNone(report_path)
        workbook = load_workbook(report_path, data_only=False)
        result_rows = list(workbook["Results"].iter_rows(values_only=True))
        result_header = {name: idx for idx, name in enumerate(result_rows[0])}
        self.assertEqual(result_rows[1][result_header["benchmark"]], "polybench/gemm")
        self.assertEqual(result_rows[1][result_header["verified"]], True)
        self.assertEqual(result_rows[1][result_header["verification_mode"]], "direct_omp")


if __name__ == "__main__":
    unittest.main()
