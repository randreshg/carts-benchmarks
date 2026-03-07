"""Excel report generation for benchmark runs."""

from __future__ import annotations

import json
import math
import re
import sys
from collections import defaultdict
from dataclasses import asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from statistics import mean, stdev
from typing import TYPE_CHECKING, Any, Dict, Iterable, List, Optional, Set, Tuple

from benchmark_common import parse_all_counters, parse_perf_csv

if TYPE_CHECKING:
    from benchmark_models import BenchmarkResult, ExperimentStep

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment
    from openpyxl.styles import Border
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:  # pragma: no cover - runtime dependency check
    Workbook = None  # type: ignore[assignment]
    CellIsRule = None  # type: ignore[assignment]
    Alignment = None  # type: ignore[assignment]
    Border = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    Side = None  # type: ignore[assignment]
    Table = None  # type: ignore[assignment]
    TableStyleInfo = None  # type: ignore[assignment]


RESULTS_COLUMNS = [
    "benchmark",
    "suite",
    "size",
    "threads",
    "nodes",
    "run",
    "run_phase",
    "compile_args",
    "profile",
    "perf_enabled",
    "perf_interval",
    "status",
    "status_detail",
    "verified",
    "has_counters",
    "has_perf",
    "verification_note",
    "verification_mode",
    "arts_checksum",
    "omp_checksum",
    "reference_checksum",
    "reference_source",
    "reference_omp_threads",
    "runtime_warning",
    "slurm_job_id",
    "slurm_state",
    "slurm_exit_code",
    "slurm_nodelist",
    "srun_error_count",
    "broken_pipe_count",
    "counter_timeout_warnings",
    "remote_send_hard_timeout_count",
    "connection_refused_count",
    "speedup_basis",
    "arts_e2e_sec",
    "omp_e2e_sec",
    "arts_kernel_sec",
    "omp_kernel_sec",
    "arts_init_sec",
    "omp_init_sec",
    "arts_total_sec",
    "omp_total_sec",
    "speedup",
    "parallel_efficiency",
    "init_overhead_pct",
    "cache_references",
    "cache_misses",
    "cache_miss_rate",
    "l1d_loads",
    "l1d_load_misses",
    "l1d_load_miss_rate",
    "counter_source",
    "counter_files_found",
    "counter_files_valid",
    "counter_expected_nodes",
    "counter_complete",
    "num_edts_created",
    "num_edts_finished",
    "num_dbs_created",
    "memory_footprint_bytes",
    "remote_bytes_sent",
    "remote_bytes_received",
    "edt_running_time_ms",
    "initialization_time_ms",
    "end_to_end_time_ms",
    "task_throughput",
    "avg_task_time_us",
    "memory_per_edt",
    "comm_bytes_per_edt",
    "artifact_run_dir",
    "artifact_run_config",
    "artifact_result_json",
    "artifact_slurm_out",
    "artifact_slurm_err",
    "artifact_build_dir",
    "artifact_arts_config",
    "artifact_counter_dir",
    "artifact_perf_dir",
    "artifact_perf_file_count",
]

SUMMARY_COLUMNS = [
    "benchmark",
    "suite",
    "size",
    "threads",
    "nodes",
    "run_phase",
    "compile_args",
    "num_runs",
    "arts_e2e_mean",
    "arts_e2e_std",
    "arts_e2e_cv_pct",
    "omp_e2e_mean",
    "omp_e2e_std",
    "arts_kernel_mean",
    "omp_kernel_mean",
    "arts_init_mean",
    "omp_init_mean",
    "speedup_mean",
    "speedup_std",
    "speedup_min",
    "speedup_max",
    "parallel_efficiency_mean",
    "init_overhead_pct_mean",
    "verified_count",
    "pass_count",
    "fail_count",
    "warn_count",
    "rows_with_counters",
    "rows_with_perf",
]

THREAD_SCALING_COLUMNS = [
    "benchmark",
    "suite",
    "size",
    "run_phase",
    "compile_args",
    "fixed_nodes",
    "baseline_threads",
    "threads",
    "num_runs",
    "arts_e2e_mean",
    "arts_e2e_std",
    "arts_e2e_cv_pct",
    "omp_e2e_mean",
    "arts_vs_omp_speedup_mean",
    "self_scaling",
    "self_scaling_efficiency",
    "verified_count",
    "pass_count",
    "rows_with_perf",
]

NODE_SCALING_COLUMNS = [
    "benchmark",
    "suite",
    "size",
    "run_phase",
    "compile_args",
    "fixed_threads",
    "baseline_nodes",
    "nodes",
    "num_runs",
    "arts_e2e_mean",
    "arts_e2e_std",
    "arts_e2e_cv_pct",
    "self_scaling",
    "self_scaling_efficiency",
    "verified_count",
    "pass_count",
    "rows_with_counters",
]

DISTRIBUTED_DB_DELTA_COLUMNS = [
    "benchmark",
    "suite",
    "size",
    "phase_family",
    "threads",
    "nodes",
    "baseline_phase",
    "distributed_phase",
    "baseline_arts_e2e_mean",
    "distributed_arts_e2e_mean",
    "time_delta_sec",
    "time_delta_pct",
    "time_ratio_dist_vs_base",
    "baseline_initialization_time_ms",
    "distributed_initialization_time_ms",
    "init_delta_pct",
    "baseline_remote_bytes_sent",
    "distributed_remote_bytes_sent",
    "remote_sent_delta_pct",
    "remote_sent_ratio_dist_vs_base",
    "baseline_remote_bytes_received",
    "distributed_remote_bytes_received",
    "remote_recv_delta_pct",
    "remote_recv_ratio_dist_vs_base",
    "baseline_memory_footprint_bytes",
    "distributed_memory_footprint_bytes",
    "memory_delta_pct",
    "memory_ratio_dist_vs_base",
    "baseline_remote_bytes_sent_cv_pct",
    "distributed_remote_bytes_sent_cv_pct",
    "remote_sent_cv_delta_pct",
    "baseline_memory_footprint_cv_pct",
    "distributed_memory_footprint_cv_pct",
    "memory_cv_delta_pct",
    "baseline_num_edts_created",
    "distributed_num_edts_created",
    "baseline_num_dbs_created",
    "distributed_num_dbs_created",
    "winner",
    "baseline_counter_dir",
    "distributed_counter_dir",
]

NODE_COUNTER_SUMMARY_COLUMNS = [
    "benchmark",
    "suite",
    "size",
    "threads",
    "nodes",
    "run",
    "run_phase",
    "compile_args",
    "counter_name",
    "capture_mode",
    "capture_level",
    "reduce_method",
    "metric_kind",
    "nodes_reported",
    "metric_total",
    "metric_mean",
    "metric_min",
    "metric_max",
    "metric_std",
    "metric_cv_pct",
    "max_to_min_ratio",
]

NODE_COUNTER_COLUMNS = [
    "benchmark",
    "suite",
    "size",
    "threads",
    "nodes",
    "run",
    "run_phase",
    "compile_args",
    "node_id",
    "total_threads",
    "counter_name",
    "capture_mode",
    "capture_level",
    "reduce_method",
    "metric_kind",
    "metric_value",
    "value",
    "value_ms",
    "history_points",
    "nodes_reported",
    "metric_total",
    "metric_mean",
    "node_fraction",
    "imbalance_vs_mean_pct",
    "counter_file",
]

PERF_FILE_COLUMNS = [
    "benchmark",
    "suite",
    "size",
    "threads",
    "nodes",
    "run",
    "run_phase",
    "compile_args",
    "perf_role",
    "perf_rank",
    "perf_file",
    "cache_references",
    "cache_misses",
    "cache_miss_rate",
    "l1d_loads",
    "l1d_load_misses",
    "l1d_load_miss_rate",
]

PATH_LIKE_FIELDS = {
    "reference_source",
    "artifact_run_dir",
    "artifact_run_config",
    "artifact_result_json",
    "artifact_slurm_out",
    "artifact_slurm_err",
    "artifact_build_dir",
    "artifact_arts_config",
    "artifact_counter_dir",
    "artifact_perf_dir",
    "counter_file",
    "perf_file",
}

COUNTER_FIELD_MAP = {
    "num_edts_created": "numEdtsCreated",
    "num_edts_finished": "numEdtsFinished",
    "num_dbs_created": "numDbsCreated",
    "memory_footprint_bytes": "memoryFootprint",
    "remote_bytes_sent": "remoteBytesSent",
    "remote_bytes_received": "remoteBytesReceived",
    "edt_running_time_ms": "edtRunningTime",
    "initialization_time_ms": "initializationTime",
    "end_to_end_time_ms": "endToEndTime",
}

INT_FIELDS = {
    "threads",
    "nodes",
    "run",
    "fixed_nodes",
    "fixed_threads",
    "baseline_threads",
    "baseline_nodes",
    "cache_references",
    "cache_misses",
    "l1d_loads",
    "l1d_load_misses",
    "num_edts_created",
    "num_edts_finished",
    "num_dbs_created",
    "srun_error_count",
    "broken_pipe_count",
    "counter_timeout_warnings",
    "remote_send_hard_timeout_count",
    "connection_refused_count",
    "counter_files_found",
    "counter_files_valid",
    "counter_expected_nodes",
    "artifact_perf_file_count",
    "memory_footprint_bytes",
    "remote_bytes_sent",
    "remote_bytes_received",
    "num_runs",
    "pass_count",
    "fail_count",
    "warn_count",
    "verified_count",
    "reference_omp_threads",
    "nodes_reported",
    "node_id",
    "total_threads",
    "history_points",
}

RATIO_FIELDS = {
    "cache_miss_rate",
    "l1d_load_miss_rate",
    "parallel_efficiency",
    "parallel_efficiency_mean",
    "self_scaling_efficiency",
    "node_fraction",
}

PCT_POINT_FIELDS = {
    "init_overhead_pct",
    "arts_e2e_cv_pct",
    "init_overhead_pct_mean",
    "imbalance_vs_mean_pct",
    "metric_cv_pct",
}


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            f = float(text)
            if math.isnan(f) or math.isinf(f):
                return None
            return f
        except ValueError:
            return None
    return None


def _safe_div(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    if numerator is None or denominator is None or denominator == 0:
        return None
    return numerator / denominator


def _mean_std(values: List[float]) -> Tuple[Optional[float], Optional[float]]:
    if not values:
        return None, None
    if len(values) == 1:
        return values[0], 0.0
    return mean(values), stdev(values)


def _geomean(values: Iterable[Optional[float]]) -> Optional[float]:
    positives = [v for v in values if v is not None and v > 0]
    if not positives:
        return None
    return math.exp(sum(math.log(v) for v in positives) / len(positives))


def _status_text(value: Any) -> str:
    raw = value.value if hasattr(value, "value") else value
    if raw is None:
        return ""
    return str(raw).upper()


def _phase_name(value: Any) -> str:
    phase = str(value or "default").strip()
    return phase or "default"


def _phase_family(value: Any) -> str:
    phase = _phase_name(value)
    for suffix in ("-baseline", "-distributed-db"):
        if phase.endswith(suffix):
            return phase[: -len(suffix)]
    return phase


def _phase_variant(phase: Any, compile_args: Any) -> Optional[str]:
    phase_name = _phase_name(phase)
    compile_text = str(compile_args or "").strip()
    if phase_name.endswith("-distributed-db") or compile_text == "--distributed-db":
        return "distributed-db"
    if phase_name.endswith("-baseline") or not compile_text:
        return "baseline"
    return None


def _counter_dir_from_artifacts(artifacts: Any) -> Optional[Path]:
    counter_dir: Optional[str] = None
    if isinstance(artifacts, dict):
        counter_dir = artifacts.get("counters_dir") or artifacts.get("counter_dir")
        if not counter_dir:
            counter_files = artifacts.get("counter_files") or []
            if counter_files:
                counter_dir = str(Path(counter_files[0]).parent)
    else:
        counter_dir = getattr(artifacts, "counters_dir", None) or getattr(artifacts, "counter_dir", None)
        if not counter_dir:
            counter_files = getattr(artifacts, "counter_files", None) or []
            if counter_files:
                counter_dir = str(Path(counter_files[0]).parent)

    if not counter_dir:
        return None
    path = Path(counter_dir)
    return path if path.exists() else None


def _counter_value(entry: Any) -> Optional[float]:
    if not isinstance(entry, dict):
        return None
    raw = entry.get("value_ms")
    if raw is None:
        raw = entry.get("value")
    return _to_float(raw)


def _count_valid_node_counter_files(counter_dir: Path) -> Tuple[int, int]:
    files = sorted(counter_dir.glob("n*.json"))
    valid = 0
    for path in files:
        try:
            with open(path) as f:
                payload = json.load(f)
        except (OSError, json.JSONDecodeError, TypeError):
            continue
        if isinstance(payload, dict):
            valid += 1
    return len(files), valid


def _aggregate_node_counter_files(counter_dir: Path) -> Tuple[Dict[str, float], int, int]:
    aggregated: Dict[str, float] = {}
    files = sorted(counter_dir.glob("n*.json"))
    valid = 0
    for path in files:
        try:
            with open(path) as f:
                payload = json.load(f)
        except (OSError, json.JSONDecodeError, TypeError):
            continue
        if not isinstance(payload, dict):
            continue
        counters = payload.get("counters")
        if not isinstance(counters, dict):
            continue

        valid += 1
        for name, entry in counters.items():
            value = _counter_value(entry)
            if value is None:
                continue
            aggregated[name] = aggregated.get(name, 0.0) + value
    return aggregated, len(files), valid


def _collect_counters(
    counter_dir: Optional[Path],
    expected_nodes: Optional[int],
) -> Tuple[Dict[str, float], Dict[str, Any]]:
    meta: Dict[str, Any] = {
        "counter_source": None,
        "counter_files_found": 0,
        "counter_files_valid": 0,
        "counter_expected_nodes": expected_nodes,
        "counter_complete": None,
    }
    if counter_dir is None:
        return {}, meta

    cluster_counters = parse_all_counters(counter_dir)
    node_counters, found, valid = _aggregate_node_counter_files(counter_dir)
    meta["counter_files_found"] = found
    meta["counter_files_valid"] = valid

    combined_counters = dict(cluster_counters)
    node_added_fields = False
    for name, value in node_counters.items():
        if name in combined_counters:
            continue
        combined_counters[name] = value
        node_added_fields = True

    if cluster_counters and node_added_fields:
        meta["counter_source"] = "cluster+node"
        if expected_nodes is not None:
            meta["counter_complete"] = valid >= expected_nodes
    elif cluster_counters:
        meta["counter_source"] = "cluster"
        if expected_nodes is not None:
            meta["counter_complete"] = True
    elif node_counters:
        meta["counter_source"] = "node_fallback"
        if expected_nodes is not None:
            meta["counter_complete"] = valid >= expected_nodes

    return combined_counters, meta


def _perf_dict(run: Any) -> Dict[str, Any]:
    if isinstance(run, dict):
        perf = run.get("perf_metrics")
    else:
        perf = getattr(run, "perf_metrics", None)

    if perf is None:
        return {}
    if isinstance(perf, dict):
        return perf
    if is_dataclass(perf):
        return asdict(perf)
    return {}


def _remap_artifact_path(path: Path, experiment_dir: Optional[Path]) -> Optional[Path]:
    if experiment_dir is None:
        return None

    experiment_name = experiment_dir.name
    parts = list(path.parts)
    if experiment_name not in parts:
        return None

    index = parts.index(experiment_name)
    suffix = parts[index + 1 :]
    candidate = experiment_dir.joinpath(*suffix)
    return candidate.resolve() if candidate.exists() else None


def _path_if_exists(value: Any, experiment_dir: Optional[Path] = None) -> Optional[Path]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    path = Path(text)
    if path.exists():
        return path.resolve()
    return _remap_artifact_path(path, experiment_dir)


def _remap_path_value(value: Any, experiment_dir: Optional[Path] = None) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    path = Path(text)
    if path.exists():
        return str(path.resolve())
    remapped = _remap_artifact_path(path, experiment_dir)
    return str(remapped) if remapped is not None else text


def _base_result_identity(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "benchmark": row.get("benchmark"),
        "suite": row.get("suite"),
        "size": row.get("size"),
        "threads": row.get("threads"),
        "nodes": row.get("nodes"),
        "run": row.get("run"),
        "run_phase": row.get("run_phase"),
        "compile_args": row.get("compile_args"),
    }


def _build_node_counter_rows(
    result_rows: List[Dict[str, Any]],
    experiment_dir: Optional[Path] = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for result_row in result_rows:
        counter_dir = _path_if_exists(
            result_row.get("artifact_counter_dir"), experiment_dir=experiment_dir
        )
        if counter_dir is None:
            continue

        for counter_file in sorted(counter_dir.glob("n*.json")):
            try:
                payload = json.loads(counter_file.read_text())
            except Exception:
                continue
            if not isinstance(payload, dict):
                continue

            metadata = payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {}
            counters = payload.get("counters")
            if not isinstance(counters, dict):
                continue

            node_id = metadata.get("nodeId")
            if node_id is None:
                match = re.match(r"n(\d+)\.json$", counter_file.name)
                node_id = int(match.group(1)) if match else None

            total_threads = metadata.get("totalThreads")

            for counter_name, entry in counters.items():
                if not isinstance(entry, dict):
                    continue

                value = _to_float(entry.get("value"))
                value_ms = _to_float(entry.get("value_ms"))
                metric_value = value_ms if value_ms is not None else value
                metric_kind = "value_ms" if value_ms is not None else "value"
                if metric_value is None:
                    continue

                history = entry.get("captureHistory")
                history_points = len(history) if isinstance(history, list) else 0

                row = _base_result_identity(result_row)
                row.update(
                    {
                        "node_id": node_id,
                        "total_threads": total_threads,
                        "counter_name": counter_name,
                        "capture_mode": entry.get("captureMode"),
                        "capture_level": entry.get("captureLevel"),
                        "reduce_method": entry.get("reduceMethod"),
                        "metric_kind": metric_kind,
                        "metric_value": metric_value,
                        "value": value,
                        "value_ms": value_ms,
                        "history_points": history_points,
                        "counter_file": str(counter_file),
                    }
                )
                rows.append(row)

    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("threads"),
            row.get("nodes"),
            row.get("run"),
            row.get("run_phase"),
            row.get("compile_args"),
            row.get("counter_name"),
            row.get("capture_mode"),
            row.get("capture_level"),
            row.get("reduce_method"),
            row.get("metric_kind"),
        )
        grouped[key].append(row)

    for grouped_rows in grouped.values():
        values = [_to_float(row.get("metric_value")) for row in grouped_rows]
        numeric_values = [value for value in values if value is not None]
        if not numeric_values:
            continue
        total_value = sum(numeric_values)
        mean_value = mean(numeric_values)
        min_value = min(numeric_values)
        max_value = max(numeric_values)
        for row in grouped_rows:
            metric_value = _to_float(row.get("metric_value"))
            row["nodes_reported"] = len(grouped_rows)
            row["metric_total"] = total_value
            row["metric_mean"] = mean_value
            row["node_fraction"] = _safe_div(metric_value, total_value)
            row["imbalance_vs_mean_pct"] = (
                ((metric_value - mean_value) / mean_value) * 100.0
                if metric_value is not None and mean_value not in (None, 0)
                else None
            )

    return rows


def _build_node_counter_summary_rows(
    result_rows: List[Dict[str, Any]],
    experiment_dir: Optional[Path] = None,
) -> List[Dict[str, Any]]:
    detail_rows = _build_node_counter_rows(result_rows, experiment_dir=experiment_dir)
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in detail_rows:
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("threads"),
            row.get("nodes"),
            row.get("run"),
            row.get("run_phase"),
            row.get("compile_args"),
            row.get("counter_name"),
            row.get("capture_mode"),
            row.get("capture_level"),
            row.get("reduce_method"),
            row.get("metric_kind"),
        )
        grouped[key].append(row)

    summary_rows: List[Dict[str, Any]] = []
    for key in sorted(grouped.keys(), key=lambda item: tuple(str(part) for part in item)):
        values = [
            _to_float(row.get("metric_value"))
            for row in grouped[key]
            if _to_float(row.get("metric_value")) is not None
        ]
        if not values:
            continue

        metric_mean, metric_std = _mean_std(values)
        metric_min = min(values)
        metric_max = max(values)
        metric_cv_pct = None
        if metric_mean not in (None, 0) and metric_std is not None:
            metric_cv_pct = (metric_std / metric_mean) * 100.0

        max_to_min_ratio = None
        if metric_min != 0:
            max_to_min_ratio = metric_max / metric_min

        (
            benchmark,
            suite,
            size,
            threads,
            nodes,
            run,
            run_phase,
            compile_args,
            counter_name,
            capture_mode,
            capture_level,
            reduce_method,
            metric_kind,
        ) = key

        summary_rows.append(
            {
                "benchmark": benchmark,
                "suite": suite,
                "size": size,
                "threads": threads,
                "nodes": nodes,
                "run": run,
                "run_phase": run_phase,
                "compile_args": compile_args,
                "counter_name": counter_name,
                "capture_mode": capture_mode,
                "capture_level": capture_level,
                "reduce_method": reduce_method,
                "metric_kind": metric_kind,
                "nodes_reported": len(values),
                "metric_total": sum(values),
                "metric_mean": metric_mean,
                "metric_min": metric_min,
                "metric_max": metric_max,
                "metric_std": metric_std,
                "metric_cv_pct": metric_cv_pct,
                "max_to_min_ratio": max_to_min_ratio,
            }
        )

    return summary_rows


def _build_perf_file_rows(
    result_rows: List[Dict[str, Any]],
    experiment_dir: Optional[Path] = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for result_row in result_rows:
        perf_dir = _path_if_exists(result_row.get("artifact_perf_dir"), experiment_dir=experiment_dir)
        if perf_dir is None:
            continue

        perf_files = sorted(perf_dir.glob("arts_node_*.csv"))
        omp_file = perf_dir / "omp.csv"
        if omp_file.exists():
            perf_files.append(omp_file)

        for perf_file in perf_files:
            perf_metrics = parse_perf_csv(perf_file)
            row = _base_result_identity(result_row)
            rank_match = re.match(r"arts_node_(\d+)\.csv$", perf_file.name)
            perf_role = "omp" if perf_file.name == "omp.csv" else "arts"
            perf_rank = int(rank_match.group(1)) if rank_match else None
            row.update(
                {
                    "perf_role": perf_role,
                    "perf_rank": perf_rank,
                    "perf_file": str(perf_file),
                    "cache_references": perf_metrics.get("cache_references") if perf_metrics else None,
                    "cache_misses": perf_metrics.get("cache_misses") if perf_metrics else None,
                    "cache_miss_rate": perf_metrics.get("cache_miss_rate") if perf_metrics else None,
                    "l1d_loads": perf_metrics.get("l1d_loads") if perf_metrics else None,
                    "l1d_load_misses": perf_metrics.get("l1d_load_misses") if perf_metrics else None,
                    "l1d_load_miss_rate": perf_metrics.get("l1d_load_miss_rate") if perf_metrics else None,
                }
            )
            rows.append(row)

    return rows


def _apply_derived_fields(row: Dict[str, Any]) -> None:
    speedup = _to_float(row.get("speedup"))
    threads = _to_float(row.get("threads"))
    arts_init = _to_float(row.get("arts_init_sec"))
    arts_e2e = _to_float(row.get("arts_e2e_sec"))

    num_edts_finished = _to_float(row.get("num_edts_finished"))
    num_edts_created = _to_float(row.get("num_edts_created"))
    end_to_end_time_ms = _to_float(row.get("end_to_end_time_ms"))
    edt_running_time_ms = _to_float(row.get("edt_running_time_ms"))
    memory_footprint_bytes = _to_float(row.get("memory_footprint_bytes"))
    remote_bytes_sent = _to_float(row.get("remote_bytes_sent"))
    remote_bytes_received = _to_float(row.get("remote_bytes_received"))

    row["parallel_efficiency"] = _safe_div(speedup, threads)

    init_overhead = _safe_div(arts_init, arts_e2e)
    row["init_overhead_pct"] = (
        init_overhead * 100.0 if init_overhead is not None else None
    )

    e2e_sec = _safe_div(end_to_end_time_ms, 1000.0)
    row["task_throughput"] = _safe_div(num_edts_finished, e2e_sec)

    avg_task_ms = _safe_div(edt_running_time_ms, num_edts_finished)
    row["avg_task_time_us"] = avg_task_ms * 1000.0 if avg_task_ms is not None else None

    row["memory_per_edt"] = _safe_div(memory_footprint_bytes, num_edts_created)

    if (
        num_edts_finished is None
        or num_edts_finished == 0
        or remote_bytes_sent is None
        or remote_bytes_received is None
    ):
        row["comm_bytes_per_edt"] = None
    else:
        row["comm_bytes_per_edt"] = (
            remote_bytes_sent + remote_bytes_received
        ) / num_edts_finished


def _verification_state(
    status: Any,
    verification_note: Any,
    arts_checksum: Any,
    omp_checksum: Any,
    reference_checksum: Any,
) -> Optional[bool]:
    status_text = _status_text(status)
    if status_text != "PASS":
        return False

    note_text = str(verification_note or "").strip().lower()
    if "mismatch" in note_text or "failed" in note_text:
        return False
    if "cannot verify" in note_text:
        return False

    if arts_checksum is not None and (omp_checksum is not None or reference_checksum is not None):
        return True

    return None


def _verification_mode_value(
    explicit_mode: Any,
    omp_checksum: Any,
    reference_checksum: Any,
) -> Optional[str]:
    if explicit_mode is not None:
        text = str(explicit_mode).strip()
        if text:
            return text
    if reference_checksum is not None:
        return "stored_omp_reference"
    if omp_checksum is not None:
        return "direct_omp"
    return None


def _empty_result_row() -> Dict[str, Any]:
    return {column: None for column in RESULTS_COLUMNS}


def _flatten_result_dataclass(result: BenchmarkResult) -> Dict[str, Any]:
    row = _empty_result_row()
    verification = result.verification
    artifacts = result.artifacts
    arts_perf_path = result.run_arts.perf_csv_path
    arts_perf_dir = str(Path(arts_perf_path).parent) if arts_perf_path else None

    row.update(
        {
            "benchmark": result.name,
            "suite": result.suite,
            "size": result.size,
            "threads": result.config.arts_threads,
            "nodes": result.config.arts_nodes,
            "run": result.run_number,
            "run_phase": _phase_name(getattr(result, "run_phase", None)),
            "compile_args": getattr(result, "compile_args", None),
            "profile": None,
            "perf_enabled": bool(result.run_arts.perf_metrics or arts_perf_path),
            "perf_interval": None,
            "status": _status_text(result.run_arts.status),
            "status_detail": _status_text(result.run_arts.status),
            "verified": _verification_state(
                result.run_arts.status,
                verification.note,
                verification.arts_checksum,
                verification.omp_checksum,
                verification.reference_checksum,
            ),
            "verification_note": verification.note,
            "verification_mode": _verification_mode_value(
                verification.mode,
                verification.omp_checksum,
                verification.reference_checksum,
            ),
            "arts_checksum": verification.arts_checksum,
            "omp_checksum": verification.omp_checksum,
            "reference_checksum": verification.reference_checksum,
            "reference_source": verification.reference_source,
            "reference_omp_threads": (
                verification.reference_omp_threads
                if verification.reference_omp_threads is not None
                else (
                    result.config.arts_threads
                    if verification.reference_checksum is not None
                    else None
                )
            ),
            "runtime_warning": False,
            "slurm_job_id": None,
            "slurm_state": None,
            "slurm_exit_code": None,
            "slurm_nodelist": None,
            "srun_error_count": None,
            "broken_pipe_count": None,
            "counter_timeout_warnings": None,
            "remote_send_hard_timeout_count": None,
            "connection_refused_count": None,
            "speedup_basis": result.timing.speedup_basis,
            "arts_e2e_sec": result.timing.arts_e2e_sec,
            "omp_e2e_sec": result.timing.omp_e2e_sec,
            "arts_kernel_sec": result.timing.arts_kernel_sec,
            "omp_kernel_sec": result.timing.omp_kernel_sec,
            "arts_init_sec": result.timing.arts_init_sec,
            "omp_init_sec": result.timing.omp_init_sec,
            "arts_total_sec": result.timing.arts_total_sec,
            "omp_total_sec": result.timing.omp_total_sec,
            "speedup": result.timing.speedup,
            "artifact_run_dir": artifacts.run_dir,
            "artifact_run_config": str(Path(artifacts.run_dir) / "run_config.json") if artifacts.run_dir else None,
            "artifact_result_json": str(Path(artifacts.run_dir) / "result.json") if artifacts.run_dir else None,
            "artifact_slurm_out": str(Path(artifacts.run_dir) / "slurm.out") if artifacts.run_dir else None,
            "artifact_slurm_err": str(Path(artifacts.run_dir) / "slurm.err") if artifacts.run_dir else None,
            "artifact_build_dir": artifacts.build_dir,
            "artifact_arts_config": artifacts.arts_config,
            "artifact_counter_dir": artifacts.counters_dir,
            "artifact_perf_dir": arts_perf_dir,
            "artifact_perf_file_count": 1 if arts_perf_path else 0,
        }
    )

    perf = _perf_dict(result.run_arts)
    row["cache_references"] = perf.get("cache_references")
    row["cache_misses"] = perf.get("cache_misses")
    row["cache_miss_rate"] = perf.get("cache_miss_rate")
    row["l1d_loads"] = perf.get("l1d_loads")
    row["l1d_load_misses"] = perf.get("l1d_load_misses")
    row["l1d_load_miss_rate"] = perf.get("l1d_load_miss_rate")

    expected_nodes = int(result.config.arts_nodes) if result.config.arts_nodes else None
    counter_dir = _counter_dir_from_artifacts(result.artifacts)
    counters, counter_meta = _collect_counters(counter_dir, expected_nodes=expected_nodes)
    for field, counter_key in COUNTER_FIELD_MAP.items():
        row[field] = counters.get(counter_key)
    row.update(counter_meta)
    row["has_counters"] = bool(counter_meta.get("counter_source"))
    row["has_perf"] = bool(result.run_arts.perf_metrics or arts_perf_path)

    _apply_derived_fields(row)
    return row


def _first_timing_value(payload: Any) -> Optional[float]:
    if isinstance(payload, dict) and payload:
        first = next(iter(payload.values()))
        return _to_float(first)
    return None


def _flatten_result_serialized(
    result: Dict[str, Any],
    experiment_dir: Optional[Path] = None,
) -> Dict[str, Any]:
    row = _empty_result_row()

    benchmark = result.get("benchmark") or result.get("name")
    config = result.get("config") or {}
    arts = result.get("arts") or result.get("run_arts") or {}
    omp = result.get("omp") or result.get("run_omp") or {}
    slurm = result.get("slurm") or {}
    diagnostics = result.get("diagnostics") or {}
    slurm_stderr = diagnostics.get("slurm_stderr") if isinstance(diagnostics, dict) else {}
    if not isinstance(slurm_stderr, dict):
        slurm_stderr = {}
    runtime_warning = diagnostics.get("runtime_warning") if isinstance(diagnostics, dict) else {}
    if not isinstance(runtime_warning, dict):
        runtime_warning = {}

    suite: Optional[str] = result.get("suite") if isinstance(result.get("suite"), str) else None
    if suite is None and isinstance(benchmark, str) and "/" in benchmark:
        suite = benchmark.split("/", 1)[0]

    status = _status_text(result.get("status") or arts.get("status"))
    status_detail = _status_text(result.get("status_detail")) or status
    detected_runtime_warning = bool(runtime_warning.get("has_warning"))
    if not detected_runtime_warning:
        detected_runtime_warning = any(
            int(slurm_stderr.get(key) or 0) > 0
            for key in (
                "srun_error_count",
                "broken_pipe_count",
                "counter_timeout_warnings",
                "remote_send_hard_timeout_count",
                "connection_refused_count",
            )
        )
    if status == "PASS" and detected_runtime_warning and status_detail == "PASS":
        status_detail = "WARN"

    row.update(
        {
            "benchmark": benchmark,
            "suite": suite,
            "size": result.get("size"),
            "threads": result.get("threads") or config.get("arts_threads"),
            "nodes": result.get("nodes") or config.get("arts_nodes"),
            "run": result.get("run_number"),
            "run_phase": _phase_name(result.get("run_phase")),
            "compile_args": result.get("compile_args"),
            "profile": result.get("profile"),
            "perf_enabled": result.get("perf"),
            "perf_interval": result.get("perf_interval"),
            "status": status,
            "status_detail": status_detail,
            "verified": _verification_state(
                status,
                (result.get("verification") or {}).get("note"),
                (result.get("verification") or {}).get("arts_checksum", arts.get("checksum")),
                (result.get("verification") or {}).get("omp_checksum", omp.get("checksum")),
                (result.get("verification") or {}).get("reference_checksum"),
            ),
            "verification_note": (result.get("verification") or {}).get("note"),
            "verification_mode": _verification_mode_value(
                (result.get("verification") or {}).get("mode"),
                (result.get("verification") or {}).get("omp_checksum", omp.get("checksum")),
                (result.get("verification") or {}).get("reference_checksum"),
            ),
            "arts_checksum": (result.get("verification") or {}).get("arts_checksum", arts.get("checksum")),
            "omp_checksum": (result.get("verification") or {}).get("omp_checksum", omp.get("checksum")),
            "reference_checksum": (result.get("verification") or {}).get("reference_checksum"),
            "reference_source": _remap_path_value(
                (result.get("verification") or {}).get("reference_source"),
                experiment_dir=experiment_dir,
            ),
            "reference_omp_threads": (
                (result.get("verification") or {}).get("reference_omp_threads")
                or (
                    result.get("threads") or config.get("arts_threads")
                    if (result.get("verification") or {}).get("reference_checksum") is not None
                    else None
                )
            ),
            "runtime_warning": detected_runtime_warning,
            "slurm_job_id": slurm.get("job_id"),
            "slurm_state": slurm.get("state"),
            "slurm_exit_code": slurm.get("exit_code"),
            "slurm_nodelist": slurm.get("nodelist"),
            "srun_error_count": slurm_stderr.get("srun_error_count"),
            "broken_pipe_count": slurm_stderr.get("broken_pipe_count"),
            "counter_timeout_warnings": slurm_stderr.get("counter_timeout_warnings"),
            "remote_send_hard_timeout_count": slurm_stderr.get("remote_send_hard_timeout_count"),
            "connection_refused_count": slurm_stderr.get("connection_refused_count"),
            "speedup_basis": None,
            "arts_e2e_sec": _to_float(arts.get("e2e_sec")) or _first_timing_value(arts.get("e2e_timings")),
            "omp_e2e_sec": _to_float(omp.get("e2e_sec")) or _first_timing_value(omp.get("e2e_timings")),
            "arts_kernel_sec": _first_timing_value(arts.get("kernel_timings")),
            "omp_kernel_sec": _first_timing_value(omp.get("kernel_timings")),
            "arts_init_sec": _to_float(arts.get("init_sec")) or _first_timing_value(arts.get("init_timings")),
            "omp_init_sec": _first_timing_value(omp.get("init_timings")),
            "arts_total_sec": _to_float(arts.get("duration_sec")),
            "omp_total_sec": _to_float(omp.get("duration_sec")),
            "speedup": _to_float(result.get("speedup")),
            "artifact_run_dir": _remap_path_value(
                (result.get("artifacts") or {}).get("run_dir"), experiment_dir=experiment_dir
            ),
            "artifact_run_config": _remap_path_value(
                (result.get("artifacts") or {}).get("run_config"), experiment_dir=experiment_dir
            ),
            "artifact_result_json": _remap_path_value(
                (result.get("artifacts") or {}).get("result_json"), experiment_dir=experiment_dir
            ),
            "artifact_slurm_out": _remap_path_value(
                (result.get("artifacts") or {}).get("slurm_out"), experiment_dir=experiment_dir
            ),
            "artifact_slurm_err": _remap_path_value(
                (result.get("artifacts") or {}).get("slurm_err"), experiment_dir=experiment_dir
            ),
            "artifact_build_dir": _remap_path_value(
                (result.get("artifacts") or {}).get("build_dir"), experiment_dir=experiment_dir
            ),
            "artifact_arts_config": _remap_path_value(
                (result.get("artifacts") or {}).get("arts_config"), experiment_dir=experiment_dir
            ),
            "artifact_counter_dir": _remap_path_value(
                (result.get("artifacts") or {}).get("counter_dir"), experiment_dir=experiment_dir
            ),
            "artifact_perf_dir": _remap_path_value(
                (result.get("artifacts") or {}).get("perf_dir"), experiment_dir=experiment_dir
            ),
        }
    )
    perf_files = (result.get("artifacts") or {}).get("perf_files")
    if isinstance(perf_files, list):
        row["artifact_perf_file_count"] = len(perf_files)
    elif row.get("artifact_perf_dir"):
        row["artifact_perf_file_count"] = 1

    perf = _perf_dict(arts)
    row["cache_references"] = perf.get("cache_references")
    row["cache_misses"] = perf.get("cache_misses")
    row["cache_miss_rate"] = perf.get("cache_miss_rate")
    row["l1d_loads"] = perf.get("l1d_loads")
    row["l1d_load_misses"] = perf.get("l1d_load_misses")
    row["l1d_load_miss_rate"] = perf.get("l1d_load_miss_rate")

    counter_dir: Optional[Path] = None
    candidate_dirs: List[Path] = []
    run_dir = result.get("_run_dir")
    if isinstance(run_dir, str) and run_dir:
        candidate_dirs.append(Path(run_dir) / "counters")

    artifact_counter_dir = _counter_dir_from_artifacts(result.get("artifacts"))
    if artifact_counter_dir is not None:
        candidate_dirs.append(artifact_counter_dir)

    if experiment_dir is not None:
        bench_name = result.get("benchmark")
        threads_value = _to_float(result.get("threads") or config.get("arts_threads"))
        nodes_value = _to_float(result.get("nodes") or config.get("arts_nodes"))
        run_number_value = _to_float(result.get("run_number"))
        if (
            isinstance(bench_name, str)
            and threads_value is not None
            and nodes_value is not None
            and run_number_value is not None
        ):
            phase = _phase_name(result.get("run_phase"))
            local_counter_dir = (
                Path(experiment_dir)
                / phase
                / bench_name
                / f"{int(threads_value)}t_{int(nodes_value)}n"
                / f"run_{int(run_number_value)}"
                / "counters"
            )
            candidate_dirs.append(local_counter_dir)

    for candidate in candidate_dirs:
        if candidate.exists():
            counter_dir = candidate
            break

    expected_nodes: Optional[int] = None
    nodes_value = _to_float(row.get("nodes"))
    if nodes_value is not None:
        expected_nodes = int(nodes_value)

    counters, counter_meta = _collect_counters(counter_dir, expected_nodes=expected_nodes)
    for field, counter_key in COUNTER_FIELD_MAP.items():
        row[field] = counters.get(counter_key)
    row.update(counter_meta)
    row["has_counters"] = bool(counter_meta.get("counter_source"))
    row["has_perf"] = bool(row.get("perf_enabled")) or bool(perf)

    _apply_derived_fields(row)
    return row


def _build_summary_rows(result_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[Any, Any, Any, Any, Any, Any, Any], List[Dict[str, Any]]] = defaultdict(list)
    for row in result_rows:
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("threads"),
            row.get("nodes"),
            row.get("run_phase"),
            row.get("compile_args"),
        )
        grouped[key].append(row)

    summary_rows: List[Dict[str, Any]] = []
    for key in sorted(
        grouped.keys(),
        key=lambda k: (
            str(k[0]),
            str(k[1]),
            str(k[2]),
            int(k[3] or 0),
            int(k[4] or 0),
            str(k[5] or ""),
            str(k[6] or ""),
        ),
    ):
        benchmark, suite, size, threads, nodes, run_phase, compile_args = key
        runs = grouped[key]

        def collect(field: str) -> List[float]:
            values = [_to_float(r.get(field)) for r in runs]
            return [v for v in values if v is not None]

        arts_e2e_values = collect("arts_e2e_sec")
        omp_e2e_values = collect("omp_e2e_sec")
        arts_kernel_values = collect("arts_kernel_sec")
        omp_kernel_values = collect("omp_kernel_sec")
        arts_init_values = collect("arts_init_sec")
        omp_init_values = collect("omp_init_sec")
        speedup_values = collect("speedup")
        efficiency_values = collect("parallel_efficiency")
        init_overhead_values = collect("init_overhead_pct")

        arts_e2e_mean, arts_e2e_std = _mean_std(arts_e2e_values)
        omp_e2e_mean, omp_e2e_std = _mean_std(omp_e2e_values)
        arts_kernel_mean, _ = _mean_std(arts_kernel_values)
        omp_kernel_mean, _ = _mean_std(omp_kernel_values)
        arts_init_mean, _ = _mean_std(arts_init_values)
        omp_init_mean, _ = _mean_std(omp_init_values)
        speedup_mean, speedup_std = _mean_std(speedup_values)
        efficiency_mean, _ = _mean_std(efficiency_values)
        init_overhead_mean, _ = _mean_std(init_overhead_values)

        if arts_e2e_mean is None or arts_e2e_std is None:
            arts_e2e_cv = None
        elif arts_e2e_mean == 0:
            arts_e2e_cv = 0.0
        else:
            arts_e2e_cv = (arts_e2e_std / arts_e2e_mean) * 100.0

        pass_count = sum(1 for r in runs if str(r.get("status", "")).upper() == "PASS")
        fail_count = len(runs) - pass_count
        warn_count = sum(1 for r in runs if r.get("runtime_warning") is True)
        verified_count = sum(1 for r in runs if r.get("verified") is True)
        rows_with_counters = sum(1 for r in runs if r.get("has_counters") is True)
        rows_with_perf = sum(1 for r in runs if r.get("has_perf") is True)

        summary_rows.append(
            {
                "benchmark": benchmark,
                "suite": suite,
                "size": size,
                "threads": threads,
                "nodes": nodes,
                "run_phase": run_phase,
                "compile_args": compile_args,
                "num_runs": len(runs),
                "arts_e2e_mean": arts_e2e_mean,
                "arts_e2e_std": arts_e2e_std,
                "arts_e2e_cv_pct": arts_e2e_cv,
                "omp_e2e_mean": omp_e2e_mean,
                "omp_e2e_std": omp_e2e_std,
                "arts_kernel_mean": arts_kernel_mean,
                "omp_kernel_mean": omp_kernel_mean,
                "arts_init_mean": arts_init_mean,
                "omp_init_mean": omp_init_mean,
                "speedup_mean": speedup_mean,
                "speedup_std": speedup_std,
                "speedup_min": min(speedup_values) if speedup_values else None,
                "speedup_max": max(speedup_values) if speedup_values else None,
                "parallel_efficiency_mean": efficiency_mean,
                "init_overhead_pct_mean": init_overhead_mean,
                "verified_count": verified_count,
                "pass_count": pass_count,
                "fail_count": fail_count,
                "warn_count": warn_count,
                "rows_with_counters": rows_with_counters,
                "rows_with_perf": rows_with_perf,
            }
        )

    phases = sorted(
        {
            _phase_name(row.get("run_phase"))
            for row in summary_rows
            if row.get("benchmark") != "GEOMEAN"
        }
    )
    for phase in phases:
        phase_rows = [
            row
            for row in summary_rows
            if row.get("benchmark") != "GEOMEAN" and _phase_name(row.get("run_phase")) == phase
        ]
        geomean_speedup = _geomean(row.get("speedup_mean") for row in phase_rows)
        if geomean_speedup is None:
            continue
        footer = {column: None for column in SUMMARY_COLUMNS}
        footer["benchmark"] = "GEOMEAN"
        footer["run_phase"] = phase
        footer["speedup_mean"] = geomean_speedup
        summary_rows.append(footer)

    return summary_rows


def _build_thread_scaling_rows(summary_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in summary_rows:
        if row.get("benchmark") == "GEOMEAN":
            continue
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("run_phase"),
            row.get("compile_args"),
            row.get("nodes"),
        )
        grouped[key].append(row)

    rows: List[Dict[str, Any]] = []
    for key in sorted(grouped.keys(), key=lambda item: tuple(str(part) for part in item)):
        configs = grouped[key]
        threads = sorted(
            {
                int(thread)
                for thread in (_to_float(row.get("threads")) for row in configs)
                if thread is not None
            }
        )
        if len(threads) <= 1:
            continue

        baseline_threads = threads[0]
        by_thread = {int(row.get("threads") or 0): row for row in configs}
        baseline_e2e = by_thread.get(baseline_threads, {}).get("arts_e2e_mean")

        for thread in threads:
            row = by_thread.get(thread)
            if row is None:
                continue
            scaling = _safe_div(_to_float(baseline_e2e), _to_float(row.get("arts_e2e_mean")))
            relative_parallelism = thread / baseline_threads if baseline_threads else None
            efficiency = (
                _safe_div(_to_float(scaling), relative_parallelism)
                if relative_parallelism not in (None, 0)
                else None
            )
            rows.append(
                {
                    "benchmark": row.get("benchmark"),
                    "suite": row.get("suite"),
                    "size": row.get("size"),
                    "run_phase": row.get("run_phase"),
                    "compile_args": row.get("compile_args"),
                    "fixed_nodes": row.get("nodes"),
                    "baseline_threads": baseline_threads,
                    "threads": thread,
                    "num_runs": row.get("num_runs"),
                    "arts_e2e_mean": row.get("arts_e2e_mean"),
                    "arts_e2e_std": row.get("arts_e2e_std"),
                    "arts_e2e_cv_pct": row.get("arts_e2e_cv_pct"),
                    "omp_e2e_mean": row.get("omp_e2e_mean"),
                    "arts_vs_omp_speedup_mean": row.get("speedup_mean"),
                    "self_scaling": scaling,
                    "self_scaling_efficiency": efficiency,
                    "verified_count": row.get("verified_count"),
                    "pass_count": row.get("pass_count"),
                    "rows_with_perf": row.get("rows_with_perf"),
                }
            )

    return rows


def _build_node_scaling_rows(summary_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in summary_rows:
        if row.get("benchmark") == "GEOMEAN":
            continue
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("run_phase"),
            row.get("compile_args"),
            row.get("threads"),
        )
        grouped[key].append(row)

    rows: List[Dict[str, Any]] = []
    for key in sorted(grouped.keys(), key=lambda item: tuple(str(part) for part in item)):
        configs = grouped[key]
        nodes = sorted(
            {
                int(node)
                for node in (_to_float(row.get("nodes")) for row in configs)
                if node is not None
            }
        )
        if len(nodes) <= 1:
            continue

        baseline_nodes = nodes[0]
        by_node = {int(row.get("nodes") or 0): row for row in configs}
        baseline_e2e = by_node.get(baseline_nodes, {}).get("arts_e2e_mean")

        for node in nodes:
            row = by_node.get(node)
            if row is None:
                continue
            scaling = _safe_div(_to_float(baseline_e2e), _to_float(row.get("arts_e2e_mean")))
            relative_parallelism = node / baseline_nodes if baseline_nodes else None
            efficiency = (
                _safe_div(_to_float(scaling), relative_parallelism)
                if relative_parallelism not in (None, 0)
                else None
            )
            rows.append(
                {
                    "benchmark": row.get("benchmark"),
                    "suite": row.get("suite"),
                    "size": row.get("size"),
                    "run_phase": row.get("run_phase"),
                    "compile_args": row.get("compile_args"),
                    "fixed_threads": row.get("threads"),
                    "baseline_nodes": baseline_nodes,
                    "nodes": node,
                    "num_runs": row.get("num_runs"),
                    "arts_e2e_mean": row.get("arts_e2e_mean"),
                    "arts_e2e_std": row.get("arts_e2e_std"),
                    "arts_e2e_cv_pct": row.get("arts_e2e_cv_pct"),
                    "self_scaling": scaling,
                    "self_scaling_efficiency": efficiency,
                    "verified_count": row.get("verified_count"),
                    "pass_count": row.get("pass_count"),
                    "rows_with_counters": row.get("rows_with_counters"),
                }
            )

    return rows


def _aggregate_node_counter_metrics(
    node_counter_summary_rows: List[Dict[str, Any]],
) -> Dict[Tuple[Any, ...], Dict[str, Any]]:
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in node_counter_summary_rows:
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("threads"),
            row.get("nodes"),
            row.get("run_phase"),
            row.get("compile_args"),
            row.get("counter_name"),
        )
        grouped[key].append(row)

    metrics: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    for key, rows in grouped.items():
        entry: Dict[str, Any] = {}
        for field in (
            "metric_total",
            "metric_mean",
            "metric_min",
            "metric_max",
            "metric_std",
            "metric_cv_pct",
            "max_to_min_ratio",
            "nodes_reported",
        ):
            values = [_to_float(row.get(field)) for row in rows]
            numeric_values = [value for value in values if value is not None]
            if not numeric_values:
                entry[field] = None
                continue
            entry[field] = mean(numeric_values)
        metrics[key] = entry

    return metrics


def _build_distributed_db_delta_rows(
    summary_rows: List[Dict[str, Any]],
    result_rows: List[Dict[str, Any]],
    node_counter_summary_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    paired_rows: Dict[Tuple[Any, ...], Dict[str, Dict[str, Any]]] = defaultdict(dict)
    for row in summary_rows:
        if row.get("benchmark") == "GEOMEAN":
            continue
        variant = _phase_variant(row.get("run_phase"), row.get("compile_args"))
        if variant not in {"baseline", "distributed-db"}:
            continue
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("threads"),
            row.get("nodes"),
            _phase_family(row.get("run_phase")),
        )
        paired_rows[key][variant] = row

    artifact_dirs: Dict[Tuple[Any, ...], str] = {}
    for row in result_rows:
        variant = _phase_variant(row.get("run_phase"), row.get("compile_args"))
        if variant not in {"baseline", "distributed-db"}:
            continue
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("threads"),
            row.get("nodes"),
            _phase_family(row.get("run_phase")),
            variant,
        )
        if key in artifact_dirs:
            continue
        counter_dir = row.get("artifact_counter_dir")
        if counter_dir:
            artifact_dirs[key] = str(counter_dir)

    aggregated_result_metrics: Dict[Tuple[Any, ...], Dict[str, Optional[float]]] = {}
    metric_fields = (
        "initialization_time_ms",
        "remote_bytes_sent",
        "remote_bytes_received",
        "memory_footprint_bytes",
        "num_edts_created",
        "num_dbs_created",
    )
    grouped_result_rows: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in result_rows:
        key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("threads"),
            row.get("nodes"),
            row.get("run_phase"),
            row.get("compile_args"),
        )
        grouped_result_rows[key].append(row)
    for key, rows in grouped_result_rows.items():
        metrics: Dict[str, Optional[float]] = {}
        for field in metric_fields:
            values = [_to_float(row.get(field)) for row in rows]
            numeric_values = [value for value in values if value is not None]
            metrics[field] = mean(numeric_values) if numeric_values else None
        aggregated_result_metrics[key] = metrics

    counter_metrics = _aggregate_node_counter_metrics(node_counter_summary_rows)

    def counter_metric(summary_row: Dict[str, Any], counter_name: str, field: str) -> Any:
        key = (
            summary_row.get("benchmark"),
            summary_row.get("suite"),
            summary_row.get("size"),
            summary_row.get("threads"),
            summary_row.get("nodes"),
            summary_row.get("run_phase"),
            summary_row.get("compile_args"),
            counter_name,
        )
        return counter_metrics.get(key, {}).get(field)

    def result_metric(summary_row: Dict[str, Any], field: str) -> Optional[float]:
        key = (
            summary_row.get("benchmark"),
            summary_row.get("suite"),
            summary_row.get("size"),
            summary_row.get("threads"),
            summary_row.get("nodes"),
            summary_row.get("run_phase"),
            summary_row.get("compile_args"),
        )
        return aggregated_result_metrics.get(key, {}).get(field)

    rows: List[Dict[str, Any]] = []
    for key in sorted(paired_rows.keys(), key=lambda item: tuple(str(part) for part in item)):
        variants = paired_rows[key]
        baseline = variants.get("baseline")
        distributed = variants.get("distributed-db")
        if baseline is None or distributed is None:
            continue

        baseline_time = _to_float(baseline.get("arts_e2e_mean"))
        distributed_time = _to_float(distributed.get("arts_e2e_mean"))
        baseline_init = result_metric(baseline, "initialization_time_ms")
        distributed_init = result_metric(distributed, "initialization_time_ms")
        baseline_sent = result_metric(baseline, "remote_bytes_sent")
        distributed_sent = result_metric(distributed, "remote_bytes_sent")
        baseline_recv = result_metric(baseline, "remote_bytes_received")
        distributed_recv = result_metric(distributed, "remote_bytes_received")
        baseline_mem = result_metric(baseline, "memory_footprint_bytes")
        distributed_mem = result_metric(distributed, "memory_footprint_bytes")
        baseline_edts = result_metric(baseline, "num_edts_created")
        distributed_edts = result_metric(distributed, "num_edts_created")
        baseline_dbs = result_metric(baseline, "num_dbs_created")
        distributed_dbs = result_metric(distributed, "num_dbs_created")

        baseline_sent_cv = _to_float(counter_metric(baseline, "remoteBytesSent", "metric_cv_pct"))
        distributed_sent_cv = _to_float(
            counter_metric(distributed, "remoteBytesSent", "metric_cv_pct")
        )
        baseline_mem_cv = _to_float(counter_metric(baseline, "memoryFootprint", "metric_cv_pct"))
        distributed_mem_cv = _to_float(
            counter_metric(distributed, "memoryFootprint", "metric_cv_pct")
        )

        time_delta_sec = (
            distributed_time - baseline_time
            if distributed_time is not None and baseline_time is not None
            else None
        )
        rows.append(
            {
                "benchmark": baseline.get("benchmark"),
                "suite": baseline.get("suite"),
                "size": baseline.get("size"),
                "phase_family": key[-1],
                "threads": baseline.get("threads"),
                "nodes": baseline.get("nodes"),
                "baseline_phase": baseline.get("run_phase"),
                "distributed_phase": distributed.get("run_phase"),
                "baseline_arts_e2e_mean": baseline.get("arts_e2e_mean"),
                "distributed_arts_e2e_mean": distributed.get("arts_e2e_mean"),
                "time_delta_sec": time_delta_sec,
                "time_delta_pct": (
                    _safe_div(time_delta_sec, baseline_time) * 100.0
                    if time_delta_sec is not None and baseline_time not in (None, 0)
                    else None
                ),
                "time_ratio_dist_vs_base": _safe_div(distributed_time, baseline_time),
                "baseline_initialization_time_ms": baseline_init,
                "distributed_initialization_time_ms": distributed_init,
                "init_delta_pct": (
                    _safe_div(distributed_init - baseline_init, baseline_init) * 100.0
                    if distributed_init is not None and baseline_init not in (None, 0)
                    else None
                ),
                "baseline_remote_bytes_sent": baseline_sent,
                "distributed_remote_bytes_sent": distributed_sent,
                "remote_sent_delta_pct": (
                    _safe_div(distributed_sent - baseline_sent, baseline_sent) * 100.0
                    if distributed_sent is not None and baseline_sent not in (None, 0)
                    else None
                ),
                "remote_sent_ratio_dist_vs_base": _safe_div(distributed_sent, baseline_sent),
                "baseline_remote_bytes_received": baseline_recv,
                "distributed_remote_bytes_received": distributed_recv,
                "remote_recv_delta_pct": (
                    _safe_div(distributed_recv - baseline_recv, baseline_recv) * 100.0
                    if distributed_recv is not None and baseline_recv not in (None, 0)
                    else None
                ),
                "remote_recv_ratio_dist_vs_base": _safe_div(distributed_recv, baseline_recv),
                "baseline_memory_footprint_bytes": baseline_mem,
                "distributed_memory_footprint_bytes": distributed_mem,
                "memory_delta_pct": (
                    _safe_div(distributed_mem - baseline_mem, baseline_mem) * 100.0
                    if distributed_mem is not None and baseline_mem not in (None, 0)
                    else None
                ),
                "memory_ratio_dist_vs_base": _safe_div(distributed_mem, baseline_mem),
                "baseline_remote_bytes_sent_cv_pct": baseline_sent_cv,
                "distributed_remote_bytes_sent_cv_pct": distributed_sent_cv,
                "remote_sent_cv_delta_pct": (
                    distributed_sent_cv - baseline_sent_cv
                    if distributed_sent_cv is not None and baseline_sent_cv is not None
                    else None
                ),
                "baseline_memory_footprint_cv_pct": baseline_mem_cv,
                "distributed_memory_footprint_cv_pct": distributed_mem_cv,
                "memory_cv_delta_pct": (
                    distributed_mem_cv - baseline_mem_cv
                    if distributed_mem_cv is not None and baseline_mem_cv is not None
                    else None
                ),
                "baseline_num_edts_created": baseline_edts,
                "distributed_num_edts_created": distributed_edts,
                "baseline_num_dbs_created": baseline_dbs,
                "distributed_num_dbs_created": distributed_dbs,
                "winner": (
                    "distributed-db"
                    if _to_float(_safe_div(distributed_time, baseline_time)) is not None
                    and _safe_div(distributed_time, baseline_time) < 1.0
                    else "baseline"
                ),
                "baseline_counter_dir": artifact_dirs.get((*key, "baseline")),
                "distributed_counter_dir": artifact_dirs.get((*key, "distributed-db")),
            }
        )

    return rows


def _build_comparison_sheet(
    workbook: Workbook,
    result_rows: List[Dict[str, Any]],
    steps: Optional[List["ExperimentStep"]] = None,
) -> None:
    all_phases = sorted({_phase_name(row.get("run_phase")) for row in result_rows})
    if len(all_phases) <= 1:
        return

    # Determine phase order: use steps definition order, fall back to sorted names.
    if steps:
        ordered_phases: List[str] = []
        for idx, step in enumerate(steps, start=1):
            name = _phase_name(getattr(step, "name", None) or f"step_{idx}")
            if name in all_phases and name not in ordered_phases:
                ordered_phases.append(name)
        for phase in all_phases:
            if phase not in ordered_phases:
                ordered_phases.append(phase)
    else:
        ordered_phases = list(all_phases)

    # Group result rows by (config_key, phase).
    counter_fields = [
        "num_edts_created",
        "num_edts_finished",
        "num_dbs_created",
        "memory_footprint_bytes",
        "remote_bytes_sent",
        "remote_bytes_received",
        "edt_running_time_ms",
        "initialization_time_ms",
        "end_to_end_time_ms",
        "task_throughput",
        "avg_task_time_us",
        "memory_per_edt",
        "comm_bytes_per_edt",
    ]
    perf_fields = [
        "cache_references",
        "cache_misses",
        "cache_miss_rate",
        "l1d_loads",
        "l1d_load_misses",
        "l1d_load_miss_rate",
    ]

    grouped: Dict[Tuple[Any, Any, Any, Any, Any, Any, str], List[Dict[str, Any]]] = defaultdict(list)
    config_keys: Set[Tuple[Any, Any, Any, Any, Any, Any]] = set()
    for row in result_rows:
        phase = _phase_name(row.get("run_phase"))
        config_key = (
            row.get("benchmark"),
            row.get("suite"),
            row.get("size"),
            row.get("threads"),
            row.get("nodes"),
            row.get("compile_args"),
        )
        config_keys.add(config_key)
        grouped[(*config_key, phase)].append(row)

    # Detect which phases have counter / perf data in any row.
    phases_with_counters: Set[str] = set()
    phases_with_perf: Set[str] = set()
    for row in result_rows:
        phase = _phase_name(row.get("run_phase"))
        if any(row.get(f) is not None for f in counter_fields):
            phases_with_counters.add(phase)
        if any(row.get(f) is not None for f in perf_fields):
            phases_with_perf.add(phase)

    # Build dynamic column list.
    columns: List[str] = ["benchmark", "suite", "size", "threads", "nodes", "compile_args"]
    for phase in ordered_phases:
        columns.extend([
            f"{phase}_arts_e2e_mean",
            f"{phase}_speedup_mean",
        ])
    for phase in ordered_phases:
        if phase in phases_with_counters:
            columns.extend(f"{phase}_{f}" for f in counter_fields)
    for phase in ordered_phases:
        if phase in phases_with_perf:
            columns.extend(f"{phase}_{f}" for f in perf_fields)

    def collect(rows: List[Dict[str, Any]], field: str) -> List[float]:
        values = [_to_float(r.get(field)) for r in rows]
        return [v for v in values if v is not None]

    ws = workbook.create_sheet(title="Comparison")
    ws.append(columns)
    _style_header(ws)
    ws.freeze_panes = "A2"

    for config_key in sorted(
        config_keys,
        key=lambda k: (
            str(k[0]),
            str(k[1]),
            str(k[2]),
            int(k[3] or 0),
            int(k[4] or 0),
            str(k[5] or ""),
        ),
    ):
        benchmark, suite, size, threads, nodes, compile_args = config_key
        row_values: List[Any] = [benchmark, suite, size, threads, nodes, compile_args]

        # Timing columns for each phase.
        for phase in ordered_phases:
            phase_rows = grouped.get((*config_key, phase), [])
            arts_e2e_mean, _ = _mean_std(collect(phase_rows, "arts_e2e_sec"))
            speedup_mean, _ = _mean_std(collect(phase_rows, "speedup"))
            row_values.extend([arts_e2e_mean, speedup_mean])

        # Counter columns for phases that have counter data.
        for phase in ordered_phases:
            if phase in phases_with_counters:
                phase_rows = grouped.get((*config_key, phase), [])
                for field in counter_fields:
                    row_values.append(_mean_std(collect(phase_rows, field))[0])

        # Perf columns for phases that have perf data.
        for phase in ordered_phases:
            if phase in phases_with_perf:
                phase_rows = grouped.get((*config_key, phase), [])
                for field in perf_fields:
                    row_values.append(_mean_std(collect(phase_rows, field))[0])

        ws.append(row_values)

    _finalize_sheet(ws, columns)


def _build_overview_sheet(
    workbook: Workbook,
    result_rows: List[Dict[str, Any]],
    steps: Optional[List["ExperimentStep"]],
    metadata: Optional[Dict[str, Any]] = None,
) -> None:
    ws = workbook.create_sheet(title="Overview")
    columns = [
        "step",
        "step_description",
        "size",
        "threads",
        "nodes",
        "runs",
        "compile_args",
        "debug",
        "perf",
        "perf_interval",
        "profile",
        "benchmarks_run",
        "rows",
        "passed",
        "verified",
        "warn",
        "failed",
        "rows_with_counters",
        "rows_with_complete_counters",
        "rows_with_perf",
    ]
    ws.append(columns)
    _style_header(ws)
    ws.freeze_panes = "A2"

    by_phase: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in result_rows:
        by_phase[_phase_name(row.get("run_phase"))].append(row)

    def append_overview_row(
        step_name: str,
        phase_rows: List[Dict[str, Any]],
        *,
        description: Any,
        size: Any,
        threads: Any,
        nodes: Any,
        runs: Any,
        compile_args: Any,
        debug: Any,
        perf: Any,
        perf_interval: Any,
        profile: Any,
        benchmarks_run: str,
    ) -> None:
        statuses = [str(r.get("status") or "").upper() for r in phase_rows]
        passed = sum(1 for s in statuses if s == "PASS")
        verified = sum(1 for r in phase_rows if r.get("verified") is True)
        failed = sum(1 for s in statuses if s in {"FAIL", "CRASH", "TIMEOUT"})
        warn = sum(1 for r in phase_rows if _status_text(r.get("status_detail")) == "WARN")
        rows_with_counters = sum(
            1
            for r in phase_rows
            if any(r.get(field) is not None for field in COUNTER_FIELD_MAP.keys())
        )
        rows_with_complete_counters = sum(1 for r in phase_rows if r.get("counter_complete") is True)
        rows_with_perf = sum(
            1
            for r in phase_rows
            if any(
                r.get(field) is not None
                for field in (
                    "cache_references",
                    "cache_misses",
                    "cache_miss_rate",
                    "l1d_loads",
                    "l1d_load_misses",
                    "l1d_load_miss_rate",
                )
            )
        )
        ws.append(
            [
                step_name,
                description,
                size,
                threads,
                nodes,
                runs,
                compile_args,
                debug,
                perf,
                perf_interval,
                profile,
                benchmarks_run,
                len(phase_rows),
                passed,
                verified,
                warn,
                failed,
                rows_with_counters,
                rows_with_complete_counters,
                rows_with_perf,
            ]
        )

    def summarize(values: List[Any]) -> Optional[str]:
        filtered = [value for value in values if value not in (None, "", False)]
        if not filtered:
            return None
        if all(isinstance(value, (int, float)) and float(value).is_integer() for value in filtered):
            normalized = sorted({int(value) for value in filtered})
            return ",".join(str(value) for value in normalized)
        normalized = sorted({str(value) for value in filtered})
        return ",".join(normalized)

    metadata_steps = None
    if isinstance(metadata, dict):
        raw_steps = metadata.get("experiment_steps")
        if isinstance(raw_steps, list):
            metadata_steps = raw_steps

    if steps or metadata_steps:
        if steps:
            step_descriptors = [
                {
                    "name": getattr(step, "name", None) or f"step_{idx}",
                    "description": getattr(step, "description", None),
                    "benchmarks": getattr(step, "benchmarks", None),
                    "size": getattr(step, "size", None),
                    "threads": getattr(step, "threads", None),
                    "nodes": getattr(step, "nodes", None),
                    "runs": getattr(step, "runs", None),
                    "compile_args": getattr(step, "compile_args", None),
                    "debug": getattr(step, "debug", None),
                    "perf": bool(getattr(step, "perf", False)),
                    "perf_interval": (
                        getattr(step, "perf_interval", None)
                        if bool(getattr(step, "perf", False))
                        else None
                    ),
                    "profile": getattr(step, "profile", None),
                }
                for idx, step in enumerate(steps, start=1)
            ]
        else:
            step_descriptors = metadata_steps

        for idx, step_desc in enumerate(step_descriptors, start=1):
            step_name = _phase_name(step_desc.get("name") or f"step_{idx}")
            phase_rows = by_phase.get(step_name, [])
            benchmarks = step_desc.get("benchmarks")
            benchmarks_run = ", ".join(benchmarks) if benchmarks else "all"
            perf_enabled = bool(step_desc.get("perf", False))
            append_overview_row(
                step_name,
                phase_rows,
                description=step_desc.get("description"),
                size=step_desc.get("size"),
                threads=step_desc.get("threads"),
                nodes=step_desc.get("nodes"),
                runs=step_desc.get("runs"),
                compile_args=step_desc.get("compile_args"),
                debug=step_desc.get("debug"),
                perf=perf_enabled,
                perf_interval=step_desc.get("perf_interval") if perf_enabled else None,
                profile=step_desc.get("profile"),
                benchmarks_run=benchmarks_run,
            )
    else:
        for step_name in sorted(by_phase.keys()):
            phase_rows = by_phase[step_name]
            append_overview_row(
                step_name,
                phase_rows,
                description=None,
                size=summarize([r.get("size") for r in phase_rows]),
                threads=summarize([r.get("threads") for r in phase_rows]),
                nodes=summarize([r.get("nodes") for r in phase_rows]),
                runs=len({r.get("run") for r in phase_rows if r.get("run") is not None}) or None,
                compile_args=summarize([r.get("compile_args") for r in phase_rows]),
                debug=None,
                perf=any(r.get("perf_enabled") for r in phase_rows),
                perf_interval=summarize([r.get("perf_interval") for r in phase_rows]),
                profile=summarize([r.get("profile") for r in phase_rows]),
                benchmarks_run=summarize([r.get("benchmark") for r in phase_rows]) or "all",
            )

    _finalize_sheet(ws, columns)
    _apply_wrapped_column_style(ws, columns, "step_description", width=56)
    _apply_wrapped_column_style(ws, columns, "benchmarks_run", width=30)
    _apply_wrapped_column_style(ws, columns, "profile", width=34)
    _style_sheet_tab(ws, "5B9BD5")


def _build_guide_sheet(
    workbook: Workbook,
    sheet_specs: List[Tuple[str, str, str]],
    *,
    experiment_name: Optional[str] = None,
    experiment_description: Optional[str] = None,
) -> None:
    ws = workbook.create_sheet(title="Guide", index=0)
    columns = ["sheet", "purpose", "when_to_use", "open"]
    ws.append(columns)
    _style_header(ws)

    for sheet_name, purpose, when_to_use in sheet_specs:
        ws.append(
            [
                sheet_name,
                purpose,
                when_to_use,
                f'=HYPERLINK("#\'{sheet_name}\'!A1","Open")',
            ]
        )

    quick_flow_row = 1
    quick_flow_col = 6
    _style_guide_summary_block(
        ws,
        title_row=quick_flow_row,
        title_col=quick_flow_col,
        title="Benchmark Report Guide",
        body_lines=[
            "Filter the tables in each sheet rather than editing raw artifacts first.",
        ],
    )
    review_start_row = quick_flow_row + 3
    if experiment_name:
        body_lines = [experiment_name]
        if experiment_description:
            body_lines.append(experiment_description)
        _style_guide_summary_block(
            ws,
            title_row=review_start_row,
            title_col=quick_flow_col,
            title="Experiment",
            body_lines=body_lines,
        )
        review_start_row += 4

    _style_guide_summary_block(
        ws,
        title_row=review_start_row,
        title_col=quick_flow_col,
        title="Suggested review order",
        body_lines=[],
    )
    quick_flow = [
        "1. Overview: confirm the planned steps and pass counts.",
        "2. Issues: check failures, warnings, and incomplete artifacts first.",
        "3. Summary: compare aggregated timings and verification coverage.",
        "4. ThreadScaling / NodeScaling / DistributedDbDelta: inspect scaling behavior.",
        "5. NodeCounterSummary / PerfFiles: inspect balance and hardware counters.",
        "6. Results: drill down to exact run artifacts when a row needs explanation.",
    ]
    for offset, text in enumerate(quick_flow, start=1):
        cell = ws.cell(row=review_start_row + offset, column=quick_flow_col, value=text)
        if PatternFill is not None:
            cell.fill = PatternFill(fill_type="solid", fgColor="F7FAFC")
        if Alignment is not None:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    _apply_table_formats(ws, columns)
    if Table is not None and TableStyleInfo is not None and len(sheet_specs) >= 1:
        table = Table(
            displayName=_excel_table_name(ws.title),
            ref=f"A1:D{len(sheet_specs) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        ws.auto_filter.ref = table.ref
    else:
        ws.auto_filter.ref = f"A1:D{len(sheet_specs) + 1}"
    _autosize_columns(ws)
    _apply_wrapped_column_style(ws, columns, "purpose", width=38)
    _apply_wrapped_column_style(ws, columns, "when_to_use", width=44)
    _set_column_width(ws, "A", 24)
    _set_column_width(ws, "D", 12)
    _set_column_width(ws, "F", 64)
    _style_sheet_tab(ws, "4472C4")


def _append_table_sheet(workbook: Workbook, title: str, columns: List[str], rows: List[Dict[str, Any]]) -> None:
    ws = workbook.create_sheet(title=title)
    ws.append(columns)
    _style_header(ws)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(column) for column in columns])

    _finalize_sheet(ws, columns)


def _build_issues_sheet(workbook: Workbook, result_rows: List[Dict[str, Any]]) -> None:
    issue_rows = [
        row
        for row in result_rows
        if (
            _status_text(row.get("status")) not in {"PASS", "SKIP"}
            or _status_text(row.get("status_detail")) == "WARN"
            or row.get("runtime_warning") is True
            or (row.get("has_counters") is True and row.get("counter_complete") is False)
            or row.get("verified") is False
        )
    ]
    if not issue_rows:
        return

    _append_table_sheet(workbook, "Issues", RESULTS_COLUMNS, issue_rows)


def _append_optional_table_sheet(
    workbook: Workbook,
    title: str,
    columns: List[str],
    rows: List[Dict[str, Any]],
) -> None:
    if not rows:
        return
    _append_table_sheet(workbook, title, columns, rows)


def _apply_path_hyperlinks(worksheet: Any, columns: List[str]) -> None:
    path_columns = [
        idx
        for idx, field in enumerate(columns, start=1)
        if field in PATH_LIKE_FIELDS
    ]
    if not path_columns or worksheet.max_row < 2:
        return

    for idx in path_columns:
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=idx)
            value = cell.value
            if not isinstance(value, str):
                continue
            text = value.strip()
            if not text:
                continue

            if re.match(r"^[A-Za-z][A-Za-z0-9+.-]*://", text):
                target = text
            else:
                path = Path(text)
                if not path.is_absolute():
                    continue
                try:
                    target = path.as_uri()
                except ValueError:
                    continue

            cell.hyperlink = target
            cell.style = "Hyperlink"


def _excel_table_name(title: str) -> str:
    base = re.sub(r"[^A-Za-z0-9_]", "_", title)
    if not base:
        base = "Sheet"
    if not re.match(r"^[A-Za-z_]", base):
        base = f"T_{base}"
    return f"tbl_{base}"[:255]


def _finalize_sheet(worksheet: Any, columns: List[str]) -> None:
    _apply_table_formats(worksheet, columns)
    _apply_speedup_rules(worksheet, columns)
    _apply_status_fill(worksheet, columns)
    _apply_path_hyperlinks(worksheet, columns)
    worksheet.auto_filter.ref = worksheet.dimensions

    if (
        Table is not None
        and TableStyleInfo is not None
        and worksheet.max_row >= 2
        and worksheet.max_column >= 1
    ):
        table = Table(
            displayName=_excel_table_name(worksheet.title),
            ref=worksheet.dimensions,
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    _autosize_columns(worksheet)


def _autosize_columns(worksheet: Any, max_width: int = 48) -> None:
    for column_cells in worksheet.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        if not values:
            continue
        width = max(len(str(v)) for v in values) + 2
        width = min(max_width, max(10, width))
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _style_header(worksheet: Any) -> None:
    if Font is None:
        return
    fill = PatternFill(fill_type="solid", fgColor="1F4E78") if PatternFill is not None else None
    border = (
        Border(bottom=Side(style="thin", color="D9E2F3"))
        if Border is not None and Side is not None
        else None
    )
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        if fill is not None:
            cell.fill = fill
        if border is not None:
            cell.border = border
        if Alignment is not None:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_column_width(worksheet: Any, column_letter: str, width: float) -> None:
    worksheet.column_dimensions[column_letter].width = width


def _apply_wrapped_column_style(
    worksheet: Any,
    columns: List[str],
    field_name: str,
    *,
    width: Optional[float] = None,
) -> None:
    if field_name not in columns:
        return
    idx = columns.index(field_name) + 1
    column_letter = worksheet.cell(row=1, column=idx).column_letter
    if width is not None:
        _set_column_width(worksheet, column_letter, width)
    if Alignment is None or worksheet.max_row < 2:
        return
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=idx)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _style_guide_summary_block(
    worksheet: Any,
    *,
    title_row: int,
    title_col: int,
    title: str,
    body_lines: List[str],
) -> None:
    worksheet.cell(row=title_row, column=title_col, value=title)
    border = (
        Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )
        if Border is not None and Side is not None
        else None
    )
    if Font is not None:
        worksheet.cell(row=title_row, column=title_col).font = Font(bold=True, size=12)
    if PatternFill is not None:
        worksheet.cell(row=title_row, column=title_col).fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )
    if border is not None:
        worksheet.cell(row=title_row, column=title_col).border = border
    if Alignment is not None:
        worksheet.cell(row=title_row, column=title_col).alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

    for offset, text in enumerate(body_lines, start=1):
        cell = worksheet.cell(row=title_row + offset, column=title_col, value=text)
        if PatternFill is not None:
            cell.fill = PatternFill(fill_type="solid", fgColor="F7FAFC")
        if border is not None:
            cell.border = border
        if Alignment is not None:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _style_sheet_tab(worksheet: Any, color: str) -> None:
    try:
        worksheet.sheet_properties.tabColor = color
    except AttributeError:
        return


TIME_SUFFIXES = {"_e2e", "_e2e_mean", "_e2e_std", "_kernel", "_kernel_mean", "_init", "_init_mean"}


def _classify_format(field: str) -> Optional[str]:
    """Return number format for a field, using suffix matching for step-prefixed columns."""
    for int_field in INT_FIELDS:
        if field == int_field or field.endswith(f"_{int_field}"):
            return "#,##0"
    for ratio_field in RATIO_FIELDS:
        if field == ratio_field or field.endswith(f"_{ratio_field}"):
            return "0.00%"
    if field == "self_scaling" or field.endswith("_self_scaling"):
        return "0.000"
    if field.endswith("_ratio") or "_ratio_" in field:
        return "0.000"
    for pct_field in PCT_POINT_FIELDS:
        if field == pct_field or field.endswith(f"_{pct_field}"):
            return "0.00"
    if field.endswith("_delta_pct"):
        return "0.00"
    for suffix in TIME_SUFFIXES:
        if field == suffix.lstrip("_") or field.endswith(suffix):
            return "0.000"
    return None


def _apply_table_formats(worksheet: Any, columns: List[str]) -> None:
    max_row = worksheet.max_row
    if max_row < 2:
        return

    for idx, field in enumerate(columns, start=1):
        fmt = _classify_format(field)
        for row in range(2, max_row + 1):
            cell = worksheet.cell(row=row, column=idx)
            value = cell.value
            if value is None or value == "":
                continue

            if fmt is not None:
                cell.number_format = fmt
            elif isinstance(value, (float, int)):
                cell.number_format = "0.000000"


def _apply_speedup_rules(worksheet: Any, columns: List[str]) -> None:
    if CellIsRule is None or PatternFill is None:
        return

    positive_good_cols = [
        idx for idx, name in enumerate(columns, start=1)
        if (
            ("speedup" in name and name not in {"peak_speedup"})
            or "self_scaling" in name
        )
    ]
    negative_good_cols = [
        idx for idx, name in enumerate(columns, start=1)
        if name.endswith("_delta_pct") or name.endswith("_ratio_dist_vs_base")
    ]
    if (not positive_good_cols and not negative_good_cols) or worksheet.max_row < 2:
        return

    green_fill = PatternFill(fill_type="solid", fgColor="E7F5E7")
    red_fill = PatternFill(fill_type="solid", fgColor="FDECEC")

    for idx in positive_good_cols:
        col = worksheet.cell(row=1, column=idx).column_letter
        rng = f"{col}2:{col}{worksheet.max_row}"
        worksheet.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=green_fill)
        )
        worksheet.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["1"], fill=red_fill)
        )

    for idx in negative_good_cols:
        col = worksheet.cell(row=1, column=idx).column_letter
        rng = f"{col}2:{col}{worksheet.max_row}"
        threshold = "0" if worksheet.cell(row=1, column=idx).value.endswith("_delta_pct") else "1"
        worksheet.conditional_formatting.add(
            rng, CellIsRule(operator="lessThanOrEqual", formula=[threshold], fill=green_fill)
        )
        worksheet.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=[threshold], fill=red_fill)
        )


def _apply_status_fill(worksheet: Any, columns: List[str]) -> None:
    if PatternFill is None:
        return
    if "status" not in columns or worksheet.max_row < 2:
        return

    idx = columns.index("status") + 1
    pass_fill = PatternFill(fill_type="solid", fgColor="E7F5E7")
    fail_fill = PatternFill(fill_type="solid", fgColor="FDECEC")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF4D6")

    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=idx)
        text = str(cell.value or "").upper()
        if not text:
            continue
        if text == "PASS":
            cell.fill = pass_fill
        elif text in {"FAIL", "CRASH", "TIMEOUT"}:
            cell.fill = fail_fill
        elif text == "SKIP":
            cell.fill = warn_fill


def _build_scaling_sheet(workbook: Workbook, summary_rows: List[Dict[str, Any]]) -> None:
    ws = workbook.create_sheet(title="ScalingMatrix")

    def _self_scaling(base_e2e: Any, arts_e2e: Any) -> Optional[float]:
        return _safe_div(_to_float(base_e2e), _to_float(arts_e2e))

    def _scaling_efficiency(scale: Any, current: int, baseline: int) -> Optional[float]:
        if baseline <= 0:
            return None
        relative_parallelism = current / baseline
        if relative_parallelism == 0:
            return None
        return _safe_div(_to_float(scale), relative_parallelism)

    data_rows = [r for r in summary_rows if r.get("benchmark") != "GEOMEAN"]
    phase_threads: Dict[str, Set[int]] = defaultdict(set)
    phase_nodes: Dict[str, Set[int]] = defaultdict(set)
    for row in data_rows:
        phase = _phase_name(row.get("run_phase"))
        threads_value = _to_float(row.get("threads"))
        nodes_value = _to_float(row.get("nodes"))
        if threads_value is not None:
            phase_threads[phase].add(int(threads_value))
        if nodes_value is not None:
            phase_nodes[phase].add(int(nodes_value))

    sweep_phases = sorted(
        {
            phase
            for phase in set(list(phase_threads.keys()) + list(phase_nodes.keys()))
            if len(phase_threads.get(phase, set())) > 1 or len(phase_nodes.get(phase, set())) > 1
        }
    )
    has_thread_sweep = any(len(values) > 1 for values in phase_threads.values())
    has_node_sweep = any(len(values) > 1 for values in phase_nodes.values())
    if sweep_phases:
        data_rows = [
            row for row in data_rows if _phase_name(row.get("run_phase")) in sweep_phases
        ]

    threads = sorted({int(r["threads"]) for r in data_rows if r.get("threads") is not None})
    nodes = sorted({int(r["nodes"]) for r in data_rows if r.get("nodes") is not None})

    if not has_thread_sweep and not has_node_sweep:
        ws["A1"] = "No thread/node sweep detected. Scaling view is empty for this run."
        _autosize_columns(ws)
        return

    indexed: Dict[Tuple[str, str, str, str, str], Dict[Tuple[int, int], Dict[str, Any]]] = defaultdict(dict)
    for row in data_rows:
        benchmark = str(row.get("benchmark") or "")
        suite = str(row.get("suite") or "")
        size = str(row.get("size") or "")
        run_phase = _phase_name(row.get("run_phase"))
        compile_args = str(row.get("compile_args") or "")
        threads_value = int(row.get("threads") or 0)
        nodes_value = int(row.get("nodes") or 0)
        indexed[(benchmark, suite, size, run_phase, compile_args)][(threads_value, nodes_value)] = row

    headers: List[str]
    matrix_rows: List[List[Any]] = []

    if has_thread_sweep and not has_node_sweep:
        headers = ["benchmark", "suite", "size", "run_phase", "compile_args"]
        for t in threads:
            headers.extend([f"T{t}_arts_e2e", f"T{t}_self_scaling", f"T{t}_efficiency"])
        headers.extend(["peak_self_scaling", "peak_threads", "max_thread_self_scaling"])

        for (benchmark, suite, size, run_phase, compile_args), configs in sorted(indexed.items()):
            row_values: List[Any] = [benchmark, suite, size, run_phase, compile_args]
            scaling_candidates: List[Tuple[float, int]] = []
            phase_threads = sorted({t for (t, _n) in configs.keys()})
            phase_nodes = sorted({n for (_t, n) in configs.keys()})
            baseline_thread = phase_threads[0] if phase_threads else threads[0]
            fixed_node = phase_nodes[0] if phase_nodes else nodes[0]
            base_e2e = configs.get((baseline_thread, fixed_node), {}).get("arts_e2e_mean")

            for t in threads:
                entry = configs.get((t, fixed_node), {})
                arts_e2e = entry.get("arts_e2e_mean")
                scaling = _self_scaling(base_e2e, arts_e2e)
                efficiency = _scaling_efficiency(scaling, t, baseline_thread)
                row_values.extend([arts_e2e, scaling, efficiency])
                scaling_f = _to_float(scaling)
                if scaling_f is not None and scaling_f > 0:
                    scaling_candidates.append((scaling_f, t))

            peak_scaling = (
                max(scaling_candidates, key=lambda x: x[0])[0] if scaling_candidates else None
            )
            peak_threads = (
                max(scaling_candidates, key=lambda x: x[0])[1] if scaling_candidates else None
            )

            max_t = max(threads)
            max_t_value = configs.get((max_t, fixed_node), {}).get("arts_e2e_mean")
            max_thread_self_scaling = _self_scaling(base_e2e, max_t_value)

            row_values.extend([peak_scaling, peak_threads, max_thread_self_scaling])
            matrix_rows.append(row_values)

    elif has_node_sweep and not has_thread_sweep:
        headers = ["benchmark", "suite", "size", "run_phase", "compile_args"]
        for n in nodes:
            headers.extend([f"N{n}_arts_e2e", f"N{n}_self_scaling", f"N{n}_efficiency"])
        headers.extend(["peak_self_scaling", "peak_nodes"])

        for (benchmark, suite, size, run_phase, compile_args), configs in sorted(indexed.items()):
            row_values = [benchmark, suite, size, run_phase, compile_args]
            scaling_candidates: List[Tuple[float, int]] = []
            phase_threads = sorted({t for (t, _n) in configs.keys()})
            phase_nodes = sorted({n for (_t, n) in configs.keys()})
            fixed_thread = phase_threads[0] if phase_threads else threads[0]
            baseline_node = phase_nodes[0] if phase_nodes else nodes[0]
            baseline_e2e = configs.get((fixed_thread, baseline_node), {}).get("arts_e2e_mean")

            for n in nodes:
                entry = configs.get((fixed_thread, n), {})
                arts_e2e = entry.get("arts_e2e_mean")
                scaling = _self_scaling(baseline_e2e, arts_e2e)
                efficiency = _scaling_efficiency(scaling, n, baseline_node)
                row_values.extend([arts_e2e, scaling, efficiency])

                scaling_f = _to_float(scaling)
                if scaling_f is not None and scaling_f > 0:
                    scaling_candidates.append((scaling_f, n))

            peak_scaling = (
                max(scaling_candidates, key=lambda x: x[0])[0] if scaling_candidates else None
            )
            peak_nodes = (
                max(scaling_candidates, key=lambda x: x[0])[1] if scaling_candidates else None
            )
            row_values.extend([peak_scaling, peak_nodes])
            matrix_rows.append(row_values)

    else:
        headers = ["benchmark", "suite", "size", "run_phase", "compile_args"]
        for t in threads:
            for n in nodes:
                prefix = f"T{t}N{n}"
                headers.extend(
                    [
                        f"{prefix}_arts_e2e",
                        f"{prefix}_self_scaling",
                        f"{prefix}_efficiency",
                    ]
                )
        headers.extend(["peak_self_scaling", "peak_threads", "peak_nodes"])

        for (benchmark, suite, size, run_phase, compile_args), configs in sorted(indexed.items()):
            row_values = [benchmark, suite, size, run_phase, compile_args]
            peak_tuple: Optional[Tuple[float, int, int]] = None
            phase_threads = sorted({t for (t, _n) in configs.keys()})
            phase_nodes = sorted({n for (_t, n) in configs.keys()})
            baseline_thread = phase_threads[0] if phase_threads else threads[0]
            baseline_node = phase_nodes[0] if phase_nodes else nodes[0]
            base_e2e = configs.get((baseline_thread, baseline_node), {}).get("arts_e2e_mean")

            for t in threads:
                for n in nodes:
                    entry = configs.get((t, n), {})
                    arts_e2e = entry.get("arts_e2e_mean")
                    scaling = _self_scaling(base_e2e, arts_e2e)
                    relative_parallelism = (t / baseline_thread) * (n / baseline_node)
                    efficiency = (
                        _safe_div(_to_float(scaling), relative_parallelism)
                        if baseline_thread > 0 and baseline_node > 0 and relative_parallelism != 0
                        else None
                    )
                    row_values.extend([arts_e2e, scaling, efficiency])

                    scaling_f = _to_float(scaling)
                    if scaling_f is not None and scaling_f > 0:
                        candidate = (scaling_f, t, n)
                        if peak_tuple is None or scaling_f > peak_tuple[0]:
                            peak_tuple = candidate

            if peak_tuple is None:
                row_values.extend([None, None, None])
            else:
                row_values.extend([peak_tuple[0], peak_tuple[1], peak_tuple[2]])

            matrix_rows.append(row_values)

    ws.append(headers)
    _style_header(ws)
    ws.freeze_panes = "A2"

    for values in matrix_rows:
        ws.append(values)

    if matrix_rows:
        phase_groups: Dict[Tuple[str, str, str], List[List[Any]]] = defaultdict(list)
        for row in matrix_rows:
            phase_groups[(str(row[2]), str(row[3]), str(row[4]))].append(row)

        for size, phase, compile_args in sorted(phase_groups.keys()):
            rows = phase_groups[(size, phase, compile_args)]
            geomean_row: List[Any] = ["GEOMEAN", "", size, phase, compile_args]
            for idx, header in enumerate(headers[5:], start=5):
                if header.startswith("peak_"):
                    geomean_row.append(None)
                    continue
                values = [_to_float(r[idx]) for r in rows]
                values = [v for v in values if v is not None and v > 0]
                geomean_row.append(_geomean(values))
            ws.append(geomean_row)

    _finalize_sheet(ws, headers)


def _load_results_metadata(experiment_dir: Path) -> Dict[str, Any]:
    results_json = experiment_dir / "results.json"
    if not results_json.exists():
        return {}

    try:
        with open(results_json) as f:
            data = json.load(f)
        return data.get("metadata", {})
    except (OSError, json.JSONDecodeError, TypeError, AttributeError):
        return {}


def _load_manifest_command(experiment_dir: Path) -> Optional[str]:
    manifest_json = experiment_dir / "manifest.json"
    if not manifest_json.exists():
        return None

    try:
        with open(manifest_json) as f:
            data = json.load(f)
        command = data.get("command")
        if command is None:
            return None
        return str(command)
    except (OSError, json.JSONDecodeError, TypeError, AttributeError):
        return None


def _metadata_rows(
    metadata: Dict[str, Any],
    command: Optional[str],
    report_summary: Optional[Dict[str, Any]] = None,
) -> List[Tuple[str, Any]]:
    rows: List[Tuple[str, Any]] = []

    def add(key: str, value: Any) -> None:
        if value is None:
            return
        if isinstance(value, (dict, list)):
            rows.append((key, json.dumps(value, sort_keys=True)))
        else:
            rows.append((key, value))

    add("timestamp", metadata.get("timestamp"))
    add("experiment_name", metadata.get("experiment_name"))
    add("experiment_description", metadata.get("experiment_description"))
    add("hostname", metadata.get("hostname"))
    add("size", metadata.get("size"))
    add("duration_seconds", metadata.get("total_duration_seconds"))
    add("runs_per_config", metadata.get("runs_per_config"))
    add("total_jobs", metadata.get("total_jobs"))
    add("submitted_jobs", metadata.get("submitted_jobs"))
    add("failed_submissions", metadata.get("failed_submissions"))
    add("partition", metadata.get("partition"))
    add("time_limit", metadata.get("time_limit"))
    add("profile", metadata.get("profile"))
    add("perf", metadata.get("perf"))
    add("perf_interval", metadata.get("perf_interval"))
    add("experiment_steps", metadata.get("experiment_steps"))

    add("thread_sweep", metadata.get("thread_sweep"))
    add("node_sweep", metadata.get("node_sweep"))
    add("fixed_threads", metadata.get("fixed_threads"))
    add("fixed_nodes", metadata.get("fixed_nodes"))
    add("launcher", metadata.get("launcher"))
    add("weak_scaling", metadata.get("weak_scaling"))

    repro = metadata.get("reproducibility") if isinstance(metadata, dict) else {}
    if isinstance(repro, dict):
        commits = repro.get("git_commits", {})
        if isinstance(commits, dict):
            add("git_carts", commits.get("carts"))
            add("git_arts", commits.get("arts"))
            add("git_carts_benchmarks", commits.get("carts_benchmarks"))

        compilers = repro.get("compilers", {})
        if isinstance(compilers, dict):
            add("compiler_clang", compilers.get("clang"))
            add("compiler_gcc", compilers.get("gcc"))

        cpu = repro.get("cpu", {})
        if isinstance(cpu, dict):
            add("cpu_model", cpu.get("model"))
            add("cpu_cores", cpu.get("cores"))
            add("cpu_physical_cores", cpu.get("physical_cores"))

    if isinstance(report_summary, dict):
        add("report_generated_at", report_summary.get("generated_at"))
        add("report_total_rows", report_summary.get("total_rows"))
        add("report_pass_count", report_summary.get("pass_count"))
        add("report_fail_count", report_summary.get("fail_count"))
        add("report_warn_count", report_summary.get("warn_count"))
        add("report_skipped_count", report_summary.get("skip_count"))
        add("report_verified_count", report_summary.get("verified_count"))
        add("report_geomean_speedup", report_summary.get("geomean_speedup"))
        add("report_rows_with_counters", report_summary.get("rows_with_counters"))
        add("report_rows_with_complete_counters", report_summary.get("rows_with_complete_counters"))
        add("report_rows_with_partial_counters", report_summary.get("rows_with_partial_counters"))
        add("report_rows_with_unknown_slurm_state", report_summary.get("rows_with_unknown_slurm_state"))
        add("report_rows_with_perf", report_summary.get("rows_with_perf"))

    add("command", command)
    return rows


def _append_metadata_sheet(
    workbook: Workbook,
    metadata: Dict[str, Any],
    command: Optional[str],
    report_summary: Optional[Dict[str, Any]] = None,
) -> None:
    ws = workbook.create_sheet(title="Metadata")
    ws.append(["key", "value"])
    _style_header(ws)

    for key, value in _metadata_rows(metadata, command, report_summary=report_summary):
        ws.append([key, value])

    ws.freeze_panes = "A2"
    _finalize_sheet(ws, ["key", "value"])
    _set_column_width(ws, "A", 28)
    _apply_wrapped_column_style(ws, ["key", "value"], "value", width=68)
    _style_sheet_tab(ws, "A5A5A5")


def _build_report_summary(result_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    statuses = [str(r.get("status") or "").upper() for r in result_rows]
    pass_count = sum(1 for s in statuses if s == "PASS")
    fail_count = sum(1 for s in statuses if s in {"FAIL", "CRASH", "TIMEOUT"})
    skip_count = sum(1 for s in statuses if s == "SKIP")
    warn_count = sum(1 for r in result_rows if r.get("runtime_warning") is True)
    verified_count = sum(1 for r in result_rows if r.get("verified") is True)

    phases = sorted({_phase_name(r.get("run_phase")) for r in result_rows})
    geomean_speedup: Dict[str, Optional[float]] = {}
    for phase in phases:
        phase_speedups = [
            _to_float(r.get("speedup"))
            for r in result_rows
            if _phase_name(r.get("run_phase")) == phase
        ]
        geomean_speedup[phase] = _geomean(
            v for v in phase_speedups if v is not None and v > 0
        )

    rows_with_counters = sum(
        1
        for r in result_rows
        if any(r.get(field) is not None for field in COUNTER_FIELD_MAP.keys())
    )
    rows_with_complete_counters = sum(1 for r in result_rows if r.get("counter_complete") is True)
    rows_with_partial_counters = sum(1 for r in result_rows if r.get("counter_source") == "node_fallback")
    rows_with_unknown_slurm_state = sum(
        1
        for r in result_rows
        if str(r.get("slurm_state") or "").upper() == "UNKNOWN"
    )
    rows_with_perf = sum(
        1
        for r in result_rows
        if r.get("has_perf") is True
    )

    return {
        "generated_at": datetime.now().isoformat(),
        "total_rows": len(result_rows),
        "pass_count": pass_count,
        "fail_count": fail_count,
        "warn_count": warn_count,
        "skip_count": skip_count,
        "verified_count": verified_count,
        "geomean_speedup": geomean_speedup,
        "rows_with_counters": rows_with_counters,
        "rows_with_complete_counters": rows_with_complete_counters,
        "rows_with_partial_counters": rows_with_partial_counters,
        "rows_with_unknown_slurm_state": rows_with_unknown_slurm_state,
        "rows_with_perf": rows_with_perf,
    }


def _write_report(
    result_rows: List[Dict[str, Any]],
    experiment_dir: Path,
    metadata: Optional[Dict[str, Any]] = None,
    command: Optional[str] = None,
    steps: Optional[List["ExperimentStep"]] = None,
) -> Optional[Path]:
    if Workbook is None:
        return None

    report_path = experiment_dir / "report.xlsx"
    metadata = dict(metadata or _load_results_metadata(experiment_dir))
    if steps:
        metadata.setdefault("experiment_name", getattr(steps[0], "_experiment_name", None))
        metadata.setdefault(
            "experiment_description",
            getattr(steps[0], "_experiment_description", None),
        )

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_rows = _build_summary_rows(result_rows)
    node_counter_summary_rows = _build_node_counter_summary_rows(
        result_rows, experiment_dir=experiment_dir
    )
    node_counter_rows = _build_node_counter_rows(result_rows, experiment_dir=experiment_dir)
    perf_file_rows = _build_perf_file_rows(result_rows, experiment_dir=experiment_dir)
    thread_scaling_rows = _build_thread_scaling_rows(summary_rows)
    node_scaling_rows = _build_node_scaling_rows(summary_rows)
    distributed_delta_rows = _build_distributed_db_delta_rows(
        summary_rows,
        result_rows,
        node_counter_summary_rows,
    )

    sheet_specs: List[Tuple[str, str, str]] = [
        (
            "Overview",
            "Step-level rollup with pass, verification, counter, and perf coverage.",
            "Start here to confirm the experiment shape and high-level outcomes.",
        )
    ]
    if any(
        _status_text(row.get("status")) not in {"PASS", "SKIP"}
        or _status_text(row.get("status_detail")) == "WARN"
        or row.get("runtime_warning") is True
        or (row.get("has_counters") is True and row.get("counter_complete") is False)
        or row.get("verified") is False
        for row in result_rows
    ):
        sheet_specs.append(
            (
                "Issues",
                "Only failures, warnings, verification problems, and incomplete artifacts.",
                "Use this sheet first when anything looks wrong.",
            )
        )
    sheet_specs.extend(
        [
            (
                "Summary",
                "Aggregated timings and coverage grouped by benchmark configuration.",
                "Use this for the main quantitative view before drilling into raw rows.",
            )
        ]
    )
    if thread_scaling_rows:
        sheet_specs.append(
            (
                "ThreadScaling",
                "Self-scaling view for single-node thread sweeps.",
                "Use this when the experiment varies thread count at a fixed node count.",
            )
        )
    if node_scaling_rows:
        sheet_specs.append(
            (
                "NodeScaling",
                "Self-scaling view for node sweeps at a fixed thread count.",
                "Use this when the experiment varies node count.",
            )
        )
    if distributed_delta_rows:
        sheet_specs.append(
            (
                "DistributedDbDelta",
                "Baseline vs --distributed-db deltas, including balance and communication counters.",
                "Use this to decide whether distributed DB helps for a given node count.",
            )
        )
    sheet_specs.extend(
        [
            (
                "ScalingMatrix",
                "Wide matrix view for mixed sweeps and geomean rollups.",
                "Use this when you want a single matrix across many thread/node points.",
            ),
            (
                "Comparison",
                "Side-by-side aggregated comparison across experiment phases.",
                "Use this to compare phases such as baseline vs distributed-db.",
            ),
        ]
    )
    if node_counter_summary_rows:
        sheet_specs.append(
            (
                "NodeCounterSummary",
                "Per-run node balance summary with min/max/std/CV statistics.",
                "Use this to spot hotspotting and imbalance quickly.",
            )
        )
        sheet_specs.append(
            (
                "NodeCounters",
                "Raw per-node counter rows for detailed distribution analysis.",
                "Use this when the node summary shows imbalance and you need the exact nodes.",
            )
        )
    if perf_file_rows:
        sheet_specs.append(
            (
                "PerfFiles",
                "Parsed perf outputs from ARTS and OMP runs.",
                "Use this for cache and L1 behavior after timing and correctness checks.",
            )
        )
    sheet_specs.extend(
        [
            (
                "Results",
                "Raw run rows with artifact paths, scheduler fields, counters, and verification.",
                "Use this to drill down to individual runs and artifacts.",
            ),
            (
                "Metadata",
                "Experiment metadata, commits, command line, and report coverage totals.",
                "Use this for reproducibility and provenance.",
            ),
        ]
    )

    _build_overview_sheet(workbook, result_rows, steps, metadata=metadata)
    _build_issues_sheet(workbook, result_rows)
    _append_table_sheet(workbook, "Summary", SUMMARY_COLUMNS, summary_rows)
    _append_optional_table_sheet(
        workbook,
        "ThreadScaling",
        THREAD_SCALING_COLUMNS,
        thread_scaling_rows,
    )
    _append_optional_table_sheet(
        workbook,
        "NodeScaling",
        NODE_SCALING_COLUMNS,
        node_scaling_rows,
    )
    _append_optional_table_sheet(
        workbook,
        "DistributedDbDelta",
        DISTRIBUTED_DB_DELTA_COLUMNS,
        distributed_delta_rows,
    )
    _build_scaling_sheet(workbook, summary_rows)
    _build_comparison_sheet(workbook, result_rows, steps=steps)
    _append_optional_table_sheet(
        workbook,
        "NodeCounterSummary",
        NODE_COUNTER_SUMMARY_COLUMNS,
        node_counter_summary_rows,
    )
    _append_optional_table_sheet(
        workbook,
        "NodeCounters",
        NODE_COUNTER_COLUMNS,
        node_counter_rows,
    )
    _append_optional_table_sheet(
        workbook,
        "PerfFiles",
        PERF_FILE_COLUMNS,
        perf_file_rows,
    )
    _append_table_sheet(workbook, "Results", RESULTS_COLUMNS, result_rows)

    effective_command = command or _load_manifest_command(experiment_dir)
    report_summary = _build_report_summary(result_rows)
    _append_metadata_sheet(
        workbook,
        metadata,
        effective_command,
        report_summary=report_summary,
    )
    experiment_name = metadata.get("experiment_name") if isinstance(metadata, dict) else None
    experiment_description = (
        metadata.get("experiment_description") if isinstance(metadata, dict) else None
    )
    _build_guide_sheet(
        workbook,
        sheet_specs,
        experiment_name=(
            str(experiment_name) if experiment_name is not None else None
        ),
        experiment_description=(
            str(experiment_description) if experiment_description is not None else None
        ),
    )

    workbook.save(report_path)
    return report_path


def generate_report(
    results: List["BenchmarkResult"],
    experiment_dir: Path,
    quiet: bool = False,
    steps: Optional[List["ExperimentStep"]] = None,
) -> Optional[Path]:
    """Generate report.xlsx from in-memory benchmark results."""
    del quiet  # kept for compatibility with caller API

    result_rows = [_flatten_result_dataclass(result) for result in results]

    command = "carts benchmarks " + " ".join(sys.argv[1:])
    return _write_report(
        result_rows,
        Path(experiment_dir),
        command=command,
        steps=steps,
    )


def generate_report_from_rows(
    result_rows: List[Dict[str, Any]],
    experiment_dir: Path,
    quiet: bool = False,
    steps: Optional[List["ExperimentStep"]] = None,
) -> Optional[Path]:
    """Generate report.xlsx from already-serialized result rows."""
    del quiet  # kept for compatibility with caller API

    normalized_rows: List[Dict[str, Any]] = []
    experiment_dir_path = Path(experiment_dir)
    for result in result_rows:
        if isinstance(result, dict) and ("arts" in result or "run_arts" in result):
            normalized_rows.append(
                _flatten_result_serialized(result, experiment_dir=experiment_dir_path)
            )
            continue

        row = _empty_result_row()
        if isinstance(result, dict):
            for key in RESULTS_COLUMNS:
                if key in result:
                    row[key] = result.get(key)
        _apply_derived_fields(row)
        normalized_rows.append(row)

    command = "carts benchmarks " + " ".join(sys.argv[1:])
    return _write_report(
        normalized_rows,
        experiment_dir_path,
        command=command,
        steps=steps,
    )
